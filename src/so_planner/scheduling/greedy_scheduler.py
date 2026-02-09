# -*- coding: utf-8 -*-
"""
Greedy-РїР»Р°РЅРёСЂРѕРІС‰РёРє РґР»СЏ so-planner (СЃРѕРІРјРµСЃС‚РёРј СЃ API, РІР°С€РёРјРё Excel Рё СЃС‚Р°СЂС‹Рј РІС‹Р·РѕРІРѕРј СЃ Session).
"""
from __future__ import annotations

import numpy as np
import argparse
import datetime as dt
import logging
from collections import defaultdict
from pathlib import Path

from typing import Any, Optional, Tuple, Iterable
from sqlalchemy import text
from sqlalchemy.orm import Session
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
from .greedy.loaders import (
    load_plan_of_sales as _L_load_plan_of_sales,
    load_bom as _L_load_bom,
    load_bom_article_name_map as _L_load_bom_article_name_map,
    load_machines as _L_load_machines,
    load_stock_any as _L_load_stock_any,
)

# Override monolith implementations with modular versions
load_plan_of_sales = _L_load_plan_of_sales
load_bom = _L_load_bom
load_machines = _L_load_machines
load_stock_any = _L_load_stock_any


# =========================
# Utils
# =========================
def _norm_col(s: str) -> str:
    return str(s).strip().lower().replace(" ", "").replace("_", "")

def _rebalance_unfixed_by_item_schedule(
    sched: pd.DataFrame,
    bom: pd.DataFrame,
    stock_map: dict | None,
    fixed_order_qty: dict[str, float] | None,
) -> pd.DataFrame:
    if sched is None or sched.empty:
        return sched

    src = sched.copy()
    df = src.copy()
    df["order_id"] = df["order_id"].astype(str)
    df["item_id"] = df["item_id"].astype(str)
    df["qty"] = pd.to_numeric(df.get("qty", 0), errors="coerce").fillna(0.0).astype(float)

    logger = logging.getLogger("so_planner.rebalance")

    # stock per item (sum across workshops)
    stock_by_item: dict[str, float] = {}
    if stock_map:
        for k, v in stock_map.items():
            if isinstance(k, tuple) and len(k) == 2:
                item = str(k[0])
            else:
                item = str(k)
            stock_by_item[item] = stock_by_item.get(item, 0.0) + float(v or 0.0)

    fixed_map: dict[str, float] = {}
    fixed_by_item: dict[str, float] = {}

    def _item_from_order_id(oid: str) -> str:
        s = str(oid or "")
        if ":" in s:
            tail = s.split(":", 1)[1]
            return tail.split("~", 1)[0]
        return s.split("-", 1)[0]

    for k, v in (fixed_order_qty or {}).items():
        try:
            fv = float(v)
        except Exception:
            continue
        if not np.isfinite(fv):
            continue
        oid = str(k)
        val = max(0.0, fv)
        fixed_map[oid] = val
        it = _item_from_order_id(oid)
        if it:
            fixed_by_item[it] = fixed_by_item.get(it, 0.0) + val

    # Build parent-child multipliers from BOM (max qty_per_parent per pair)
    pair_map: dict[tuple[str, str], float] = {}
    if bom is not None and not bom.empty:
        for r in bom.itertuples(index=False):
            parent = str(getattr(r, "root_item_id", "") or "")
            child = str(getattr(r, "item_id", "") or "")
            if not parent or not child or parent == child:
                continue
            mult = float(getattr(r, "qty_per_parent", 1.0) or 1.0)
            if mult <= 0:
                mult = 1.0
            key = (parent, child)
            prev = pair_map.get(key)
            if prev is None or mult > prev:
                pair_map[key] = mult

    if not pair_map:
        return df

    parents_by_child: dict[str, list[tuple[str, float]]] = {}
    children: dict[str, set[str]] = {}
    indeg: dict[str, int] = {}
    nodes: set[str] = set()
    for (parent, child), mult in pair_map.items():
        parents_by_child.setdefault(child, []).append((parent, float(mult)))
        children.setdefault(parent, set()).add(child)
        nodes.add(parent)
        nodes.add(child)
        indeg[child] = indeg.get(child, 0) + 1
        indeg.setdefault(parent, indeg.get(parent, 0))

    queue = sorted([n for n in nodes if indeg.get(n, 0) == 0])
    topo: list[str] = []
    while queue:
        n = queue.pop(0)
        topo.append(n)
        for ch in sorted(children.get(n, set())):
            indeg[ch] = indeg.get(ch, 0) - 1
            if indeg[ch] == 0:
                queue.append(ch)
                queue.sort()
    for n in sorted(nodes):
        if n not in topo:
            topo.append(n)

    # order qty per order_id, item_id (max)
    order_qty_df = df.groupby(["order_id", "item_id"], as_index=False)["qty"].max()
    order_qty: dict[str, float] = {}
    order_item: dict[str, str] = {}
    orders_by_item: dict[str, list[str]] = {}
    for r in order_qty_df.itertuples(index=False):
        oid = str(r.order_id)
        item = str(r.item_id)
        order_item[oid] = item
        if oid in fixed_map:
            qty_val = fixed_map.get(oid, 0.0)
        else:
            qty_val = float(getattr(r, "qty", 0.0) or 0.0)
        if not np.isfinite(qty_val):
            qty_val = 0.0
        order_qty[oid] = float(qty_val or 0.0)
        orders_by_item.setdefault(item, []).append(oid)
    existing_order_ids: set[str] = set(order_qty.keys())

    def _round_allocate(oids: list[str], target: float) -> dict[str, int]:
        target_int = int(round(target))
        if target_int < 0:
            target_int = 0
        raw = []
        for oid in oids:
            old_qty = float(order_qty.get(oid, 0.0) or 0.0)
            raw_qty = old_qty
            base = int(raw_qty // 1)
            rem = raw_qty - base
            raw.append({"oid": oid, "base": base, "rem": rem})
        sum_base = sum(r["base"] for r in raw)
        delta = target_int - sum_base
        if raw:
            if delta > 0:
                raw.sort(key=lambda r: r["rem"], reverse=True)
                for i in range(delta):
                    raw[i % len(raw)]["base"] += 1
            elif delta < 0:
                raw.sort(key=lambda r: r["rem"])
                i = 0
                steps = -delta
                while steps > 0 and raw:
                    r = raw[i % len(raw)]
                    if r["base"] > 0:
                        r["base"] -= 1
                        steps -= 1
                    i += 1
        return {r["oid"]: int(r["base"]) for r in raw}

    def _unique_rebalance_oid(seed: str) -> str:
        i = 1
        while True:
            cand = f"{seed}~reb{i}"
            if cand not in existing_order_ids:
                existing_order_ids.add(cand)
                return cand
            i += 1

    def _spawn_unfixed_order(item_id: str, qty_target: float) -> tuple[str | None, float, pd.DataFrame | None]:
        if qty_target <= 0:
            return None, 0.0, None
        item_oids = list(orders_by_item.get(item_id, []) or [])
        if not item_oids:
            return None, 0.0, None
        # Prefer cloning a fixed branch for this item to preserve route/date shape.
        template_oid = next((x for x in item_oids if x in fixed_map), item_oids[0])
        tpl = df[df["order_id"] == template_oid].copy()
        if tpl.empty:
            return None, 0.0, None
        new_qty = int(round(float(qty_target)))
        if new_qty <= 0:
            return None, 0.0, None
        new_oid = _unique_rebalance_oid(template_oid)
        tpl["order_id"] = new_oid
        tpl["qty"] = float(new_qty)
        return new_oid, float(new_qty), tpl

    # scale unfixed orders per item to match demand
    for item in topo:
        if item not in parents_by_child:
            continue
        demand = 0.0
        for parent, mult in parents_by_child.get(item, []):
            for oid in orders_by_item.get(parent, []):
                parent_qty = order_qty.get(oid, 0.0)
                demand += float(parent_qty or 0.0) * float(mult or 1.0)

        # Keep fixed demand by item even if exact fixed order_id is absent in current expanded schedule
        # (e.g. branch renumbering with ~ suffix on repeat run).
        fixed_sum = float(fixed_by_item.get(str(item), 0.0) or 0.0)
        unfixed_oids: list[str] = []
        unfixed_sum = 0.0
        for oid in orders_by_item.get(item, []):
            qty = float(order_qty.get(oid, 0.0) or 0.0)
            if oid in fixed_map:
                continue
            else:
                unfixed_sum += qty
                unfixed_oids.append(oid)

        stock_qty = float(stock_by_item.get(item, 0.0) or 0.0)
        target_unfixed = demand - stock_qty - fixed_sum
        if not np.isfinite(target_unfixed):
            target_unfixed = 0.0
        if target_unfixed < 0:
            target_unfixed = 0.0

        if not unfixed_oids and target_unfixed > 0:
            new_oid, new_qty, new_rows = _spawn_unfixed_order(str(item), target_unfixed)
            if new_oid and new_rows is not None:
                df = pd.concat([df, new_rows], ignore_index=True)
                order_item[new_oid] = str(item)
                order_qty[new_oid] = float(new_qty)
                orders_by_item.setdefault(str(item), []).append(new_oid)
                unfixed_oids = [new_oid]
                unfixed_sum = float(new_qty)
                logger.info(
                    "rebalance item=%s spawned_unfixed_order=%s qty=%.3f",
                    item, new_oid, new_qty,
                )

        if not unfixed_oids:
            continue

        if unfixed_sum <= 0:
            rounded = _round_allocate(unfixed_oids, target_unfixed)
            for oid, new_qty in rounded.items():
                order_qty[oid] = float(new_qty)
                df.loc[df["order_id"] == oid, "qty"] = float(new_qty)
            logger.info(
                "rebalance item=%s demand=%.3f stock=%.3f fixed=%.3f unfixed=%.3f -> %.3f (rounded=%d)",
                item, demand, stock_qty, fixed_sum, unfixed_sum, target_unfixed, int(round(target_unfixed)),
            )
            continue

        factor = target_unfixed / unfixed_sum if unfixed_sum else 1.0
        if not np.isfinite(factor):
            factor = 1.0
        if abs(factor - 1.0) < 1e-9:
            continue

        # scale (float) then round per order with largest remainder
        for oid in unfixed_oids:
            order_qty[oid] = float(order_qty.get(oid, 0.0) or 0.0) * factor
        rounded = _round_allocate(unfixed_oids, target_unfixed)
        for oid, new_qty in rounded.items():
            order_qty[oid] = float(new_qty)
            df.loc[df["order_id"] == oid, "qty"] = float(new_qty)

        logger.info(
            "rebalance item=%s demand=%.3f stock=%.3f fixed=%.3f unfixed=%.3f -> %.3f (rounded=%d)",
            item, demand, stock_qty, fixed_sum, unfixed_sum, target_unfixed, int(round(target_unfixed)),
        )

    df = df[pd.to_numeric(df.get("qty", 0), errors="coerce").fillna(0.0) > 0].copy()
    if df.empty and not src.empty:
        logger.warning("rebalance validation: result is empty, fallback to original schedule")
        return src
    return df


def _as_date(x: Any):
    try:
        return pd.to_datetime(x, errors="coerce").date()
    except Exception:
        return pd.NaT


def _ensure_positive_int(x: Any) -> int:
    try:
        v = float(x)
        if pd.isna(v) or v <= 0:
            return 0
        return int(round(v))
    except Exception:
        return 0


# =========================
# Loaders
# =========================
def _dead_load_plan_of_sales(path: Path) -> pd.DataFrame:
    return _L_load_plan_of_sales(path)
    """
    РћР¶РёРґР°РµРј wide-С‚Р°Р±Р»РёС†Сѓ:
      - РєРѕР»РѕРЅРєР° Р°СЂС‚РёРєСѓР»Р°: 'article' (РёР»Рё item_id/РњР°С‚РµСЂРёР°Р» Рё С‚.Рї.)
      - РѕСЃС‚Р°Р»СЊРЅС‹Рµ РєРѕР»РѕРЅРєРё вЂ” РґР°С‚С‹; Р·РЅР°С‡РµРЅРёСЏ вЂ” qty
    """
    df = pd.read_excel(path, sheet_name=0, dtype=object)
    norm = {_norm_col(c): c for c in df.columns}

    # РќР°С…РѕРґРёРј РєРѕР»РѕРЅРєСѓ Р°СЂС‚РёРєСѓР»Р°
    id_candidates = ["article", "item_id", "item", "РјР°С‚РµСЂРёР°Р»", "Р°СЂС‚РёРєСѓР»"]
    item_col = None
    for c in id_candidates:
        if _norm_col(c) in norm:
            item_col = norm[_norm_col(c)]
            break
    if item_col is None:
        # fallback вЂ” РїРµСЂРІР°СЏ РєРѕР»РѕРЅРєР° СЃС‡РёС‚Р°РµРј Р°СЂС‚РёРєСѓР»РѕРј
        item_col = df.columns[0]

    # Р”Р°С‚Р°-РєРѕР»РѕРЅРєРё вЂ” С‚Рµ, С‡С‚Рѕ РїР°СЂСЃСЏС‚СЃСЏ РІ РґР°С‚Сѓ
    date_cols = []
    for c in df.columns:
        if c == item_col:
            continue
        d = _as_date(c)
        if pd.notna(d):
            date_cols.append(c)

    if not date_cols:
        raise ValueError(
            f"Р’ РїР»Р°РЅРµ РЅРµ РЅР°Р№РґРµРЅС‹ РєРѕР»РѕРЅРєРё-РґР°С‚С‹. РќР°С€Р»РёСЃСЊ: {list(df.columns)}"
        )

    long_df = df.melt(
        id_vars=[item_col],
        value_vars=date_cols,
        var_name="due_date",
        value_name="qty",
    )
    long_df["due_date"] = pd.to_datetime(long_df["due_date"], errors="coerce").dt.date
    long_df.rename(columns={item_col: "item_id"}, inplace=True)
    long_df["item_id"] = long_df["item_id"].astype(str).str.strip()
    long_df["qty"] = long_df["qty"].apply(_ensure_positive_int)
    long_df = long_df[
        (long_df["qty"] > 0)
        & long_df["item_id"].ne("")
        & long_df["due_date"].notna()
    ].copy()
    long_df["priority"] = pd.to_datetime(long_df["due_date"])
    long_df.sort_values(["priority", "item_id"], inplace=True, kind="stable")
    long_df.reset_index(drop=True, inplace=True)
    return long_df


def _legacy_load_machines(path: Path) -> pd.DataFrame:
    # ... СЃСѓС‰РµСЃС‚РІСѓСЋС‰РёР№ РєРѕРґ РІС‹С€Рµ ...
    # РєР°Р»РµРЅРґР°СЂСЊ (РѕРїС†РёРѕРЅР°Р»СЊРЅРѕ)
    # ...

    # --- РќРћР’РћР•: overload_pct РЅР° СѓСЂРѕРІРЅРµ РјР°С€РёРЅС‹ (РґРѕР»СЏ; 0.25 = +25%)
    overload_cols = ["overload_pct", "overload pct", "overload%", "РїРµСЂРµРіСЂСѓР·РєР°", "РїРµСЂРµРіСЂСѓР·РєР°%"]
    df["overload_pct"] = 0.0
    for oc in overload_cols:
        if _norm_col(oc) in norm:
            s = pd.to_numeric(df[norm[_norm_col(oc)]], errors="coerce").fillna(0.0).astype(float)
            # РµСЃР»Рё РєС‚Рѕ-С‚Рѕ РґР°Р» РїСЂРѕС†РµРЅС‚С‹ РІ 0..100 вЂ” РїРµСЂРµРІРµРґС‘Рј РІ РґРѕР»СЋ
            df["overload_pct"] = np.where(s > 1.0, s / 100.0, s)
            break

    keep = ["machine_id", "capacity_per_day", "overload_pct"]
    if "calendar_date" in df.columns:
        keep.append("calendar_date")
    if "capacity_override" in df.columns:
        keep.append("capacity_override")
    return df[keep]


def _dead_load_bom(path: Path) -> pd.DataFrame:
    return _L_load_bom(path)
    """
    РџРѕРґРґРµСЂР¶РёРІР°РµС‚:
    B) РІР°С€Сѓ СЃС…РµРјСѓ:
       - article
       - operations (С€Р°Рі)
       - machine id
       - machine time (С‡Р°СЃС‹/РµРґ) РР›Р human time (С‡Р°СЃС‹/РµРґ)
       - setting time (С‡Р°СЃС‹ РЅР° РѕРїРµСЂР°С†РёСЋ, РѕРїС†РёРѕРЅР°Р»СЊРЅРѕ)
       - root article (РёРµСЂР°СЂС…РёСЏ)
    A) РєР»Р°СЃСЃРёС‡РµСЃРєСѓСЋ:
       - item_id / step / machine_id / time_per_unit (РјРёРЅ/РµРґ) [+ setup*] [+ root article?]

    Р’РѕР·РІСЂР°С‰Р°РµС‚: item_id, step, machine_id, time_per_unit (РјРёРЅ/РµРґ), setup_minutes (РјРёРЅ/РѕРї), root_item_id
    """
    df = pd.read_excel(path, sheet_name=0, dtype=object)
    norm = {_norm_col(c): c for c in df.columns}

    def has(x: str) -> bool:
        return _norm_col(x) in norm

    def col(x: str) -> str:
        return norm[_norm_col(x)]

    # --- РЎС…РµРјР° B (РІР°С€Рё С„Р°Р№Р»С‹)
    if has("article") and (has("machineid") or has("machine id")):
        out = pd.DataFrame()
        out["item_id"] = df[col("article")].astype(str).str.strip()
        out["step"] = pd.to_numeric(df[col("operations")], errors="coerce").fillna(1).astype(int) if has("operations") else 1
        mid_col = col("machineid") if has("machineid") else col("machine id")
        out["machine_id"] = df[mid_col].astype(str).str.strip()
        # РІСЂРµРјСЏ РЅР° РµРґ. (С‡Р°СЃС‹ -> РјРёРЅСѓС‚С‹)
        if has("machinetime"):
            h = pd.to_numeric(df[col("machinetime")], errors="coerce").fillna(0).astype(float)
        elif has("humantime"):
            h = pd.to_numeric(df[col("humantime")], errors="coerce").fillna(0).astype(float)
        else:
            raise ValueError("BOM: РЅРµС‚ 'machine time' РёР»Рё 'human time' (С‡Р°СЃС‹/РµРґ).")
        out["time_per_unit"] = h * 60.0
        # РЅР°Р»Р°РґРєР° (С‡Р°СЃС‹ -> РјРёРЅСѓС‚С‹)
        if has("settingtime") or has("setting time"):
            sc = col("settingtime") if has("settingtime") else col("setting time")
            out["setup_minutes"] = pd.to_numeric(df[sc], errors="coerce").fillna(0).astype(float) * 60.0
        else:
            out["setup_minutes"] = 0.0
        # root
        if has("rootarticle") or has("root article"):
            rac = col("rootarticle") if has("rootarticle") else col("root article")
            out["root_item_id"] = df[rac].astype(str).str.strip().replace({"nan": "", "None": ""})
        else:
            out["root_item_id"] = ""

        # optional workshop
        if has("workshop"):
            out["workshop"] = df[col("workshop")].astype(str).str.strip()
        else:
            out["workshop"] = ""
        # optional qty_per_parent
        if has("qty_per_parent") or has("qtyperparent"):
            qcol = col("qty_per_parent") if has("qty_per_parent") else col("qtyperparent")
            out["qty_per_parent"] = pd.to_numeric(df[qcol], errors="coerce").fillna(1.0).astype(float)
        else:
            out["qty_per_parent"] = 1.0

        out = out.sort_values(["item_id", "step"], kind="stable").reset_index(drop=True)
        ret_cols = ["item_id","step","machine_id","time_per_unit","setup_minutes","root_item_id"]
        if "workshop" in out.columns: ret_cols.append("workshop")
        if "qty_per_parent" in out.columns: ret_cols.append("qty_per_parent")
        return out[ret_cols]

    # --- РЎС…РµРјР° A (РєР»Р°СЃСЃРёРєР°)
    rename_map = {}
    candidates = {
        "item_id": ["item_id", "item", "article", "РјР°С‚РµСЂРёР°Р»", "Р°СЂС‚РёРєСѓР»"],
        "step": ["step", "operations", "operationseq", "opseq", "seq", "sequence", "РїРѕСЂСЏРґРѕРє"],
        "machine_id": ["machine_id", "machine", "resource", "СЃС‚Р°РЅРѕРє", "РјР°С€РёРЅР°", "machine id"],
        "time_per_unit": ["time_per_unit", "proc_time", "duration", "minutes_per_unit", "РјРёРЅ_РЅР°_РµРґ", "РјРёРЅСѓС‚РЅР°РµРґ", "machine time"],
    }
    for k, opts in candidates.items():
        for o in opts:
            if _norm_col(o) in norm:
                rename_map[norm[_norm_col(o)]] = k
                break
    df = df.rename(columns=rename_map)

    need = {"item_id", "machine_id", "time_per_unit"}
    missing = [c for c in need if c not in df.columns]
    if missing:
        raise ValueError(f"BOM: РѕС‚СЃСѓС‚СЃС‚РІСѓСЋС‚ {missing}. РќР°С€Р»РёСЃСЊ: {list(df.columns)}")

    if "step" not in df.columns:
        df["step"] = 1

    df["item_id"] = df["item_id"].astype(str).str.strip()
    df["step"] = pd.to_numeric(df["step"], errors="coerce").fillna(1).astype(int)
    df["machine_id"] = df["machine_id"].astype(str).str.strip()
    df["machine_id"] = df["machine_id"] \
        .str.replace(r"\.0$", "", regex=True) \
        .str.replace(r"\s+", " ", regex=True)
    
    df["time_per_unit"] = pd.to_numeric(df["time_per_unit"], errors="coerce").fillna(0).astype(float)
    if df["time_per_unit"].median() < 1.0:
        df["time_per_unit"] *= 60.0  # С‡Р°СЃС‹ -> РјРёРЅСѓС‚С‹

    df["machine_id"] = df["machine_id"].astype(str).str.strip()
    df["machine_id"] = df["machine_id"] \
        .str.replace(r"\.0$", "", regex=True) \
        .str.replace(r"\s+", " ", regex=True)
    

    # setup (РѕРїС†.)
    setup_series = None
    for cand in ["setup_minutes", "setting_time", "setup", "РЅР°Р»Р°РґРєР°", "setting time"]:
        if _norm_col(cand) in norm:
            setup_series = pd.to_numeric(df[norm[_norm_col(cand)]], errors="coerce").fillna(0).astype(float)
            if setup_series.median() < 1.0:
                setup_series *= 60.0
            break
    df["setup_minutes"] = setup_series if setup_series is not None else 0.0

    # root
    if _norm_col("root article") in norm or _norm_col("rootarticle") in norm:
        key = norm[_norm_col("root article")] if _norm_col("root article") in norm else norm[_norm_col("rootarticle")]
        df["root_item_id"] = df[key].astype(str).str.strip().replace({"nan": "", "None": ""})
    else:
        df["root_item_id"] = ""

    # optional A-schema workshop/qty_per_parent
    if _norm_col("workshop") in norm:
        df["workshop"] = df[norm[_norm_col("workshop")]].astype(str).str.strip()
    else:
        df["workshop"] = ""
    if _norm_col("qty_per_parent") in norm or _norm_col("qtyperparent") in norm:
        key = norm[_norm_col("qty_per_parent")] if _norm_col("qty_per_parent") in norm else norm[_norm_col("qtyperparent")]
        df["qty_per_parent"] = pd.to_numeric(df[key], errors="coerce").fillna(1.0).astype(float)
    else:
        df["qty_per_parent"] = 1.0

    df = df.sort_values(["item_id", "step"], kind="stable").reset_index(drop=True)
    ret_cols = ["item_id","step","machine_id","time_per_unit","setup_minutes","root_item_id"]
    if "workshop" in df.columns: ret_cols.append("workshop")
    if "qty_per_parent" in df.columns: ret_cols.append("qty_per_parent")
    return df[ret_cols]


def _dead_load_machines(path: Path) -> pd.DataFrame:
    return _L_load_machines(path)
    df = pd.read_excel(path, sheet_name=0, dtype=object)
    norm = {_norm_col(c): c for c in df.columns}

    def has(x: str) -> bool: return _norm_col(x) in norm
    def col(x: str) -> str:  return norm[_norm_col(x)]

    # machine_id
    if has("machine_id"): df = df.rename(columns={col("machine_id"): "machine_id"})
    elif has("machineid"): df = df.rename(columns={col("machineid"): "machine_id"})
    elif has("machine id"): df = df.rename(columns={col("machine id"): "machine_id"})
    else: raise ValueError("machines: РЅРµС‚ РєРѕР»РѕРЅРєРё machine id / machine_id.")

    df["machine_id"] = df["machine_id"].astype(str).str.strip()
    df["machine_id"] = df["machine_id"] \
        .str.replace(r"\.0$", "", regex=True) \
        .str.replace(r"\s+", " ", regex=True)    
    

    # capacity_per_day
    # РµСЃР»Рё РµСЃС‚СЊ capacity_per_day (РІ С‡Р°СЃР°С…/РґРµРЅСЊ) вЂ” РїРµСЂРµРІРµРґС‘Рј РІ РјРёРЅСѓС‚С‹
    if "capacity_per_day" in df.columns:
        df["capacity_per_day"] = pd.to_numeric(df["capacity_per_day"], errors="coerce").fillna(0.0)
        df["capacity_per_day"] = df["capacity_per_day"] * 60.0
    elif "count" in df.columns and ("available time" in df.columns or "available_time" in df.columns):
        at_col = "available time" if "available time" in df.columns else "available_time"
        cnt = pd.to_numeric(df["count"], errors="coerce").fillna(0.0)
        hrs = pd.to_numeric(df[at_col], errors="coerce").fillna(0.0)
        df["capacity_per_day"] = (cnt * hrs * 60.0)
    else:
        raise ValueError("machines: РЅРµ РЅР°Р№РґРµРЅС‹ РїРѕР»СЏ РґР»СЏ СЂР°СЃС‡С‘С‚Р° РјРѕС‰РЅРѕСЃС‚Рё")

    if "capacity_override" in df.columns:
        df["capacity_override"] = pd.to_numeric(df["capacity_override"], errors="coerce").fillna(pd.NA)

    # РєР°Р»РµРЅРґР°СЂСЊ (РѕРїС†.)
    if has("calendar_date") or has("date"):
        c = col("calendar_date") if has("calendar_date") else col("date")
        df = df.rename(columns={c: "calendar_date"})
        df["calendar_date"] = pd.to_datetime(df["calendar_date"], errors="coerce").dt.date
    if has("capacity_override") or has("override"):
        c = col("capacity_override") if has("capacity_override") else col("override")
        df = df.rename(columns={c: "capacity_override"})
        df["capacity_override"] = pd.to_numeric(df["capacity_override"], errors="coerce").astype(float)

    # overload_pct (РґРѕР»СЏ; РјРѕР¶РЅРѕ 0..100 в†’ РїРµСЂРµРІРµРґС‘Рј)
    df["overload_pct"] = 0.0
    for oc in ["overload_pct", "overload pct", "overload%", "РїРµСЂРµРіСЂСѓР·РєР°", "РїРµСЂРµРіСЂСѓР·РєР°%"]:
        if has(oc):
            s = pd.to_numeric(df[col(oc)], errors="coerce").fillna(0.0).astype(float)
            df["overload_pct"] = np.where(s > 1.0, s / 100.0, s)
            break

    keep = ["machine_id", "capacity_per_day", "overload_pct"]
    if "calendar_date" in df.columns: keep.append("calendar_date")
    if "capacity_override" in df.columns: keep.append("capacity_override")
    return df[keep]


def _dead_load_stock_any(path: Path) -> pd.DataFrame:
    return _L_load_stock_any(path)
    """
    Р—Р°РіСЂСѓР¶Р°РµС‚ Excel СЃ РѕСЃС‚Р°С‚РєР°РјРё, РіРёР±РєРѕ СЂР°СЃРїРѕР·РЅР°РІР°СЏ РЅР°Р·РІР°РЅРёСЏ РєРѕР»РѕРЅРѕРє.
    РџРѕРґРґРµСЂР¶РёРІР°РµРјС‹Рµ СЃРёРЅРѕРЅРёРјС‹:
      - РєР»СЋС‡ Р°СЂС‚РёРєСѓР»Р°: 'item_id','item','article','РјР°С‚РµСЂРёР°Р»','Р°СЂС‚РёРєСѓР»'
      - РєРѕР»РёС‡РµСЃС‚РІРѕ: 'stock_qty','qty','quantity','РѕСЃС‚Р°С‚РѕРє','СЃРІРѕР±РѕРґРЅС‹Р№РѕСЃС‚Р°С‚РѕРє','free_stock','on_hand','available'
    Р’РѕР·РІСЂР°С‰Р°РµС‚ df СЃ РєРѕР»РѕРЅРєР°РјРё: item_id, stock_qty (float), Р°РіСЂРµРіРёСЂРѕРІР°РЅРѕ РїРѕ item_id.
    """
    df = pd.read_excel(path, sheet_name=0, dtype=object)
    norm = {_norm_col(c): c for c in df.columns}

    key_opts = ["item_id","item","article","РјР°С‚РµСЂРёР°Р»","Р°СЂС‚РёРєСѓР»"]
    qty_opts = ["stock_qty","qty","quantity","РѕСЃС‚Р°С‚РѕРє","СЃРІРѕР±РѕРґРЅС‹Р№РѕСЃС‚Р°С‚РѕРє","free_stock","on_hand","available"]

    key_col = None
    for k in key_opts:
        if _norm_col(k) in norm:
            key_col = norm[_norm_col(k)]
            break
    if key_col is None:
        key_col = df.columns[0]

    qty_col = None
    for q in qty_opts:
        if _norm_col(q) in norm:
            qty_col = norm[_norm_col(q)]
            break
    if qty_col is None:
        raise ValueError("stock: РЅРµ РЅР°Р№РґРµРЅР° РєРѕР»РѕРЅРєР° РєРѕР»РёС‡РµСЃС‚РІР° (РЅР°РїСЂРёРјРµСЂ, 'stock_qty'/'qty'/'РѕСЃС‚Р°С‚РѕРє').")

    out = pd.DataFrame({
        "item_id": df[key_col].astype(str).str.strip(),
        "stock_qty": pd.to_numeric(df[qty_col], errors="coerce").fillna(0.0).astype(float),
    })
    out = out[out["item_id"].ne("")].copy()
    out = out.groupby("item_id", as_index=False)["stock_qty"].sum()
    return out

# --- Support tables (stock snapshot + order metadata) ---

def _ensure_support_tables(db: Session) -> None:
    """Ensure auxiliary tables used by reports and stock snapshots exist (SQLite)."""
    stmts = [
        """
        CREATE TABLE IF NOT EXISTS stock_snapshot (
            id INTEGER PRIMARY KEY,
            name TEXT,
            taken_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            notes TEXT
        )
        """,
        """
        CREATE TABLE IF NOT EXISTS stock_line (
            id INTEGER PRIMARY KEY,
            snapshot_id INTEGER NOT NULL,
            item_id TEXT NOT NULL,
            workshop TEXT DEFAULT '',
            stock_qty INTEGER NOT NULL
        )
        """,
        """
        CREATE INDEX IF NOT EXISTS ix_stock_line_main
        ON stock_line(snapshot_id,item_id,workshop)
        """,
        """
        CREATE TABLE IF NOT EXISTS plan_order_info (
            plan_id INTEGER,
            order_id TEXT,
            due_date DATE,
            PRIMARY KEY (plan_id, order_id)
        )
        """,
    ]
    for sql in stmts:
        db.execute(text(sql))
    db.commit()


def compute_orders_timeline(sched: pd.DataFrame) -> pd.DataFrame:
    """
    РўР°Р№РјР»Р°Р№РЅ РїРѕ РєР°Р¶РґРѕРјСѓ order_id:
      start_date, finish_date, duration_days, due_date, finish_lag.
    РћР¶РёРґР°РµРјС‹Рµ РєРѕР»РѕРЅРєРё РІ sched: order_id, item_id, date, due_date.
    """
    if sched.empty:
        return pd.DataFrame(columns=[
            "order_id","item_id","start_date","finish_date","duration_days","due_date","finish_lag"
        ])

    # СѓР±РµР¶РґР°РµРјСЃСЏ РІ РєРѕСЂСЂРµРєС‚РЅС‹С… С‚РёРїР°С… РґР°С‚
    df = sched.copy()
    df["date"] = pd.to_datetime(df["date"])
    df["due_date"] = pd.to_datetime(df["due_date"])

    g = df.groupby("order_id", as_index=False).agg(
        start_date=("date", "min"),
        finish_date=("date", "max"),
        due_date=("due_date", "max"),
        item_id=("item_id", "first"),
    )
    g["duration_days"] = (g["finish_date"] - g["start_date"]).dt.days + 1
    g["finish_lag"] = (g["finish_date"] - g["due_date"]).dt.days
    # РїСЂРёРІРѕРґРёРј Рє date
    g["start_date"] = g["start_date"].dt.date
    g["finish_date"] = g["finish_date"].dt.date
    g["due_date"] = g["due_date"].dt.date

    return g.sort_values(
        ["start_date","finish_date","order_id"], kind="stable"
    ).reset_index(drop=True)


def compute_order_items_timeline(sched: pd.DataFrame) -> pd.DataFrame:
    """
    РўР°Р№РјР»Р°Р№РЅ РїРѕ РєР°Р¶РґРѕР№ РїР°СЂРµ (order_id, item_id):
      start_date, finish_date, duration_days, due_date, finish_lag.
    РџРѕР»РµР·РЅРѕ, РµСЃР»Рё РІРєР»СЋС‡С‘РЅ split_child_orders Рё order_id = "<base>:<item>".
    """
    if sched.empty:
        return pd.DataFrame(columns=[
            "order_id","item_id","start_date","finish_date","duration_days","due_date","finish_lag"
        ])

    df = sched.copy()
    df["date"] = pd.to_datetime(df["date"])
    df["due_date"] = pd.to_datetime(df["due_date"])

    g = df.groupby(["order_id","item_id"], as_index=False).agg(
        start_date=("date", "min"),
        finish_date=("date", "max"),
        due_date=("due_date", "max"),
    )
    g["duration_days"] = (g["finish_date"] - g["start_date"]).dt.days + 1
    g["finish_lag"] = (g["finish_date"] - g["due_date"]).dt.days
    g["start_date"] = g["start_date"].dt.date
    g["finish_date"] = g["finish_date"].dt.date
    g["due_date"] = g["due_date"].dt.date

    return g.sort_values(
        ["order_id","item_id","start_date"], kind="stable"
    ).reset_index(drop=True)


# =========================
# Demand / order_id
# =========================
def build_demand(plan_df: pd.DataFrame) -> pd.DataFrame:
    # РѕР¶РёРґР°РЅРёРµ: РґР»РёРЅРЅС‹Р№ С„РѕСЂРјР°С‚ СЃ РєРѕР»РѕРЅРєР°РјРё item_id, due_date, qty
    # РµСЃР»Рё РїСЂРёС€С‘Р» В«С€РёСЂРѕРєРёР№В», РїСЂРѕРіРѕРЅРёС‚Рµ С‡РµСЂРµР· load_plan_of_sales Р”Рћ РІС‹Р·РѕРІР° build_demand (СЃРј. run_pipeline РЅРёР¶Рµ)
    g = plan_df.groupby(["item_id", "due_date"], as_index=False).agg(qty=("qty", "sum"))
    g = g.sort_values(["due_date", "item_id"]).reset_index(drop=True)

    # СЃС‚Р°Р±РёР»СЊРЅС‹Р№ order_id: (item_id, due_date, seq)
    from collections import defaultdict
    seq = defaultdict(int)
    order_ids = []
    for _, r in g.iterrows():
        key = (r["item_id"], r["due_date"])
        seq[key] += 1
        oid = f'{r["item_id"]}-{pd.to_datetime(r["due_date"]).strftime("%Y%m%d")}-{seq[key]:04d}'
        order_ids.append(oid)
    g["order_id"] = order_ids
    g["priority"] = pd.to_datetime(g["due_date"])  # РІР°Р¶РЅРѕ: СЂРµР°Р»СЊРЅР°СЏ РґР°С‚Р°, Р° РЅРµ 0
    return g[["order_id", "item_id", "due_date", "qty", "priority"]]


# =========================
# root->child
# =========================


def build_bom_hierarchy(bom: pd.DataFrame) -> pd.DataFrame:
    """
    Р’РѕР·РІСЂР°С‰Р°РµС‚ С‚Р°Р±Р»РёС†Сѓ СЃСЃС‹Р»РѕРє root->child Рё СѓСЂРѕРІРµРЅСЊ (0=РєРѕСЂРµРЅСЊ РІРµСЂС…РЅРµРіРѕ СѓСЂРѕРІРЅСЏ).
    РЈС‡РёС‚С‹РІР°РµС‚ С‚РѕР»СЊРєРѕ РїР°СЂС‹, РіРґРµ root_item_id РЅРµРїСѓСЃС‚РѕР№.
    """
    links = bom[["item_id","root_item_id"]].drop_duplicates()
    links = links[links["root_item_id"].fillna("").astype(str).str.len() > 0].copy()
    if links.empty:
        links["level"] = 0
        return links

    # РѕС†РµРЅРёРј СѓСЂРѕРІРЅРё СЂРµРєСѓСЂСЃРёРІРЅРѕ (РїСЂРѕСЃС‚Р°СЏ С‚РѕРїРѕР»РѕРіРёСЏ РїРѕ С†РµРїРѕС‡РєР°Рј)
    parents = {r.item_id: r.root_item_id for r in links.itertuples(index=False)}
    level = {}
    def lvl(x, seen=None):
        seen = seen or set()
        if x in level: return level[x]
        p = parents.get(x, "")
        if not p or p == x or p in seen:
            level[x] = 0
            return 0
        seen.add(x)
        level[x] = 1 + lvl(p, seen)
        return level[x]

    items = set(links["item_id"]).union(set(links["root_item_id"]))
    for it in items:
        lvl(it)

    # СѓСЂРѕРІРµРЅСЊ child = lvl(child), РєРѕСЂРµРЅСЊ РІС‹С€Рµ РїРѕ С‡РёСЃР»Сѓ
    links["level"] = links["item_id"].map(level).fillna(0).astype(int)
    return links.sort_values(["level","root_item_id","item_id"]).reset_index(drop=True)

# Р’РІРµСЂС…Сѓ С„Р°Р№Р»Р° СЂСЏРґРѕРј СЃ РёРјРїРѕСЂС‚Р°РјРё
# ============================================================================


def expand_demand_with_hierarchy(
    demand: pd.DataFrame,
    bom: pd.DataFrame,
    *,
    split_child_orders: bool = False,
    include_parents: bool = False,
    reserved_order_ids: Iterable[str] | None = None,
    fixed_order_qty: dict[str, float] | None = None,
) -> pd.DataFrame:
    # Build parent and children maps from BOM
    parents: dict[str, str] = {}
    children_map: dict[str, dict[str, float]] = {}
    for r in bom.itertuples(index=False):
        p = r.root_item_id
        c = r.item_id
        parents[c] = p
        if p and p != c:
            children_map.setdefault(p, {})[c] = float(getattr(r, "qty_per_parent", 1.0)) or 1.0

    # Р’СЃРїРѕРјРѕРіР°С‚РµР»СЊРЅС‹Рµ РѕР±С…РѕРґС‹
    def ancestors(x: str) -> list[str]:
        out = []
        seen = set()
        cur = x
        for _ in range(100000):
            p = parents.get(cur, "")
            if not p or p in seen or p == cur:
                break
            out.append(p)
            seen.add(p)
            cur = p
        return out

    def descendants_with_factor(x: str) -> list[tuple[str, float, str]]:
        """
        Р’РѕР·РІСЂР°С‰Р°РµРј РІСЃРµС… РїРѕС‚РѕРјРєРѕРІ СЃ РЅР°РєРѕРїР»РµРЅРЅС‹Рј РєРѕСЌС„С„РёС†РёРµРЅС‚РѕРј.
        Р“Р»РѕР±Р°Р»СЊРЅРѕ РЅРµ РґРµРґСѓРїРёРј, С‡С‚РѕР±С‹ СЂРµР±С‘РЅРѕРє РјРѕРі РїСЂРёР№С‚Рё РїРѕ РЅРµСЃРєРѕР»СЊРєРёРј СЂРѕРґРёС‚РµР»СЊСЃРєРёРј РІРµС‚РєР°Рј;
        С†РёРєР» РѕР±СЂС‹РІР°РµРј С‚РѕР»СЊРєРѕ РїРѕ С‚РµРєСѓС‰РµРјСѓ РїСѓС‚Рё.
        """
        out = []
        stack = [(x, 1.0, {x})]  # cur, factor, path
        while stack:
            cur, f, path = stack.pop()
            for ch, r in (children_map.get(cur, {}) or {}).items():
                if ch in path:
                    continue
                f_new = f * (r if np.isfinite(r) and r > 0 else 1.0)
                out.append((ch, f_new, cur))
                stack.append((ch, f_new, path | {ch}))
        return out

    rows = []
    for r in demand.itertuples(index=False):
        base_oid = str(r.order_id)
        it = str(r.item_id)
        due = r.due_date
        qty = int(r.qty)
        pr = r.priority

        # СЃР°Рј СЃРїСЂРѕСЃ (FG)
        rows.append(dict(
            base_order_id=base_oid,
            order_id=(f"{base_oid}:{it}" if split_child_orders else base_oid),
            item_id=it, due_date=due, qty=qty, priority=pr, role="FG"
        ))
        # РїСЂРµРґРєРё (Р±РµР· РјР°СЃС€С‚Р°Р±РёСЂРѕРІР°РЅРёСЏ)
        for a in ancestors(it):
            rows.append(dict(
                base_order_id=base_oid,
                order_id=(f"{base_oid}:{a}" if split_child_orders else base_oid),
                item_id=a, due_date=due, qty=qty, priority=pr, role="PARENT"
            ))
        # РїРѕС‚РѕРјРєРё (РјР°СЃС€С‚Р°Р±РёСЂСѓРµРј РІРЅРёР·)
        for d, fmul, parent in descendants_with_factor(it):
            rows.append(dict(
                base_order_id=base_oid,
                order_id=(f"{base_oid}:{d}" if split_child_orders else base_oid),
                item_id=d, due_date=due, qty=int(round(qty * float(fmul))), priority=pr, role="CHILD",
                parent_item_id=str(parent)
            ))

    exp = pd.DataFrame(rows)
    if exp.empty:
        raise ValueError("Greedy: expanded_demand is empty вЂ” РїСЂРѕРІРµСЂСЊС‚Рµ BOM/РїР»Р°РЅ.")
    # depth РґР»СЏ СЃРѕСЂС‚РёСЂРѕРІРєРё: СЃС‚СЂРѕРёРј СѓСЂРѕРІРЅРё РёР· BOM
    links = build_bom_hierarchy(bom)
    depth_map = {r.item_id: int(r.level) for r in links.itertuples(index=False)} if not links.empty else {}
    exp["depth"] = exp["item_id"].map(depth_map).fillna(0).astype(int)
    # РЎС‚Р°Р±РёР»СЊРЅР°СЏ СЃРѕСЂС‚РёСЂРѕРІРєР°: СЃРЅР°С‡Р°Р»Р° РїСЂРёРѕСЂРёС‚РµС‚, Р·Р°С‚РµРј depth
    exp = exp.sort_values(["priority","depth","item_id"], kind="stable").reset_index(drop=True)
    return exp

def greedy_schedule(
    demand: pd.DataFrame,
    bom: pd.DataFrame,
    machines: pd.DataFrame,
    start_date: dt.date | None = None,
    overload_pct: float = 0.0,
    split_child_orders: bool = True,
    align_roots_to_due: bool = True,
    guard_limit_days: int = 200 * 365,
    stock_map: dict | None = None,
    include_parents: bool = False,
    reserved_order_ids: Iterable[str] | None = None,
    fixed_order_qty: dict[str, float] | None = None,
    expand: bool = True,
) -> pd.DataFrame:
    """
    РџР»Р°РЅРёСЂРѕРІС‰РёРє СЃ РјР°СЂС€СЂСѓС‚Р°РјРё (step), РЅР°Р»Р°РґРєРѕР№, РёРµСЂР°СЂС…РёРµР№ BOM Рё РїРµСЂРµРіСЂСѓР·РєРѕР№.

    Р РµР¶РёРјС‹:
      - ASAP (align_roots_to_due=False): РїР»Р°РЅРёСЂСѓРµРј РІРїРµСЂС‘Рґ РѕС‚ start_date.
      - JIT  (align_roots_to_due=True): РєРѕСЂРЅРµРІРѕР№ item (role='FG') Р·Р°РєР°РЅС‡РёРІР°РµС‚ СЂРѕРІРЅРѕ РІ due_date,
        РµРіРѕ РїРѕС‚РѕРјРєРё РїР»Р°РЅРёСЂСѓСЋС‚СЃСЏ РќРђР—РђР” С‚Р°Рє, С‡С‚РѕР±С‹ Р·Р°РєРѕРЅС‡РёС‚СЊ РЅРµ РїРѕР·Р¶Рµ СЃС‚Р°СЂС‚Р° СЂРѕРґРёС‚РµР»СЏ, Рё РЅРµ СЃС‚Р°СЂС‚РѕРІР°С‚СЊ СЂР°РЅРµРµ 'today'.

    Р’ СЂР°СЃРїРёСЃР°РЅРёРµ РґРѕР±Р°РІР»СЏРµС‚СЃСЏ 'base_order_id' (РµСЃР»Рё split_child_orders=True).
    """
    
    if start_date is None:
        start_date = dt.date.today()
    warnings: list[str] = []   # в†ђ Р±СѓРґРµРј СЃРѕР±РёСЂР°С‚СЊ РїСЂРµРґСѓРїСЂРµР¶РґРµРЅРёСЏ РґР»СЏ UI/Р»РѕРіРѕРІ

    # --- map item -> workshop (from BOM); РёСЃРїРѕР»СЊР·СѓРµС‚СЃСЏ РґР»СЏ СЃРєР»Р°РґРѕРІ Рё РјР°СЂС€СЂСѓС‚РёР·Р°С†РёРё
    item_workshop = {}
    if "workshop" in bom.columns:
        try:
            _bw = bom[["item_id","workshop"]].dropna()
            _bw["item_id"] = _bw["item_id"].astype(str).str.strip()
            _bw["workshop"] = _bw["workshop"].astype(str).str.strip()
            item_workshop = _bw.drop_duplicates("item_id").set_index("item_id")["workshop"].to_dict()
        except Exception:
            item_workshop = {}
    else:
        item_workshop = {}

    use_level_demand = True

    # Р•СЃР»Рё РµСЃС‚СЊ СЃРєР»Р°Рґ, СЃРЅР°С‡Р°Р»Р° РіР°СЃРёРј СЃРїСЂРѕСЃ РїРѕ СЂРѕРґРёС‚РµР»СЋ РґРѕ СЂР°Р·РІРѕСЂРѕС‚Р° BOM,
    # С‡С‚РѕР±С‹ РЅРµ РїР»РѕРґРёС‚СЊ РґРѕС‡РµСЂРЅРёРµ Р·Р°РєР°Р·С‹ РЅР° РѕР±СЉС‘РјС‹, СѓР¶Рµ Р·Р°РєСЂС‹С‚С‹Рµ РіРѕС‚РѕРІРѕР№ РїСЂРѕРґСѓРєС†РёРµР№.
    if expand and stock_map and not use_level_demand:
        smap_prefilter = {}
        for k, v in stock_map.items():
            if isinstance(k, tuple) and len(k) == 2:
                smap_prefilter[(str(k[0]), str(k[1]))] = float(v or 0.0)
            else:
                smap_prefilter[str(k)] = float(v or 0.0)

        pre_rows = []
        demand_sorted = demand.sort_values(["priority","item_id"], kind="stable").reset_index(drop=True)
        demand_cols = list(demand.columns)
        for r in demand_sorted.itertuples(index=False):
            q = int(getattr(r, "qty", 0) or 0)
            if q <= 0:
                continue
            item = str(getattr(r, "item_id"))
            wk = item_workshop.get(item, "")
            needed = q
            tried = []
            for key in ((item, wk), (item, ""), item):
                if key in tried:
                    continue
                tried.append(key)
                avail = smap_prefilter.get(key, 0.0)
                if avail <= 0:
                    continue
                take = min(needed, int(avail))
                smap_prefilter[key] = avail - take
                needed -= take
                if needed <= 0:
                    break
            if needed > 0:
                row = {c: getattr(r, c) for c in demand_cols if hasattr(r, c)}
                if "workshop" in demand_cols and (row.get("workshop") in (None, "", pd.NA)):
                    row["workshop"] = wk
                row["qty"] = int(needed)
                pre_rows.append(row)
        demand = pd.DataFrame(pre_rows, columns=demand_cols) if pre_rows else demand.iloc[0:0]
        stock_map = smap_prefilter  # РѕСЃС‚Р°С‚РѕРє РїРѕР№РґС‘С‚ РЅРёР¶Рµ (РЅР° РґРѕС‡РµСЂРЅРёРµ СѓСЂРѕРІРЅРё)

    # РљР°СЂС‚Р° СЂРѕРґРёС‚РµР»СЏ РїРѕ BOM (РЅСѓР¶РЅР°, С‡С‚РѕР±С‹ РіР°СЃРёС‚СЊ РІРµС‚РєРё, РµСЃР»Рё СЂРѕРґРёС‚РµР»СЊ Р·Р°РєСЂС‹С‚ СЃРєР»Р°РґРѕРј)
    parent_lookup: dict[str, set[str]] = {}
    if "root_item_id" in bom.columns:
        try:
            tmp = bom[["item_id", "root_item_id"]].dropna()
            for r in tmp.itertuples(index=False):
                ch = str(r.item_id)
                par = str(r.root_item_id)
                if par and par != ch:
                    parent_lookup.setdefault(ch, set()).add(par)
        except Exception:
            parent_lookup = {}

    # === 0) РџРѕРґРіРѕС‚РѕРІРєР° РІС…РѕРґРЅС‹С… РґР°РЅРЅС‹С… Рё РёРµСЂР°СЂС…РёРё ===
    if expand:
        demand = expand_demand_with_hierarchy(
            demand,
            bom,
            split_child_orders=split_child_orders,
            include_parents=include_parents,
            reserved_order_ids=reserved_order_ids,
            fixed_order_qty=fixed_order_qty,
            stock_map=stock_map,
            item_workshop=item_workshop,
        )
        if use_level_demand:
            stock_map = None
    print("[GREEDY DEBUG] expanded_demand rows:", len(demand))
    # === STOCK CONSUMPTION (apply to all levels FG/PARENT/CHILD with proportional scaling) ===
    if stock_map and not use_level_demand:
        # normalize stock map keys to tuple (item, workshop) or str item for legacy
        smap = {}
        for k, v in stock_map.items():
            if isinstance(k, tuple) and len(k) == 2:
                smap[(str(k[0]), str(k[1]))] = float(v or 0.0)
            else:
                smap[str(k)] = float(v or 0.0)

        # scale_product: СЃРєРѕР»СЊРєРѕ РґРѕР»РµР№ РёСЃС…РѕРґРЅРѕРіРѕ qty СЂРµР°Р»СЊРЅРѕ РёРґС‘С‚ РІ РїСЂРѕРёР·РІРѕРґСЃС‚РІРѕ РґР»СЏ РєР°Р¶РґРѕРіРѕ item (per base_order_id)
        scale_product = defaultdict(dict)  # base_oid -> {item_id -> fraction_of_original}
        adjusted_map: dict[tuple[object, ...], dict] = {}
        demand = demand.sort_values(["priority","depth","item_id"], kind="stable").reset_index(drop=True)
        for r in demand.itertuples(index=False):
            item = str(r.item_id)
            wk = item_workshop.get(item, "")
            base_oid = str(getattr(r, "base_order_id", r.order_id))
            q_orig = float(r.qty) if not pd.isna(r.qty) else 0.0
            if q_orig <= 0:
                continue

            par = str(getattr(r, "parent_item_id", "") or "")
            if not par:
                parents_set = parent_lookup.get(item, set())
                # РµСЃР»Рё РЅРµСЃРєРѕР»СЊРєРѕ СЂРѕРґРёС‚РµР»РµР№, Р±РµСЂС‘Рј max РґРѕР»СЋ РІС‹РїСѓСЃРєР° СЃСЂРµРґРё РЅРёС… (Р°РіСЂРµРіРёСЂСѓРµРј РЅРёР¶Рµ)
                parent_scale = max((scale_product[base_oid].get(p, 1.0) for p in parents_set), default=1.0)
            else:
                parent_scale = scale_product[base_oid].get(par, 1.0) if par else 1.0
            required = q_orig * parent_scale
            if required <= 0:
                scale_product[base_oid][item] = 0.0
                continue

            # try to consume stock: exact (item, wk) -> generic (item, "") -> legacy item -> any other workshop for item
            needed = required
            tried_keys = []
            for key in ((item, wk), (item, ""), item):
                if key in tried_keys:
                    continue
                tried_keys.append(key)
                avail = smap.get(key, 0.0)
                if avail <= 0:
                    continue
                take = min(needed, float(avail))
                smap[key] = avail - take
                needed -= take
                if needed <= 0:
                    break

            # Fallback: sweep other workshops for this item if nothing or not enough taken (covers BOM without workshop)
            if needed > 0:
                for key, avail in list(smap.items()):
                    if isinstance(key, tuple) and len(key) == 2 and str(key[0]) == item and key not in tried_keys:
                        if avail <= 0:
                            continue
                        take = min(needed, float(avail))
                        smap[key] = avail - take
                        needed -= take
                        if needed <= 0:
                            break

            produced = max(needed, 0.0)
            # produced СЃРµР№С‡Р°СЃ = РѕСЃС‚Р°С‚РѕРє РїРѕСЃР»Рµ СЃРїРёСЃР°РЅРёСЏ СЃРєР»Р°РґР°; СЃС‡РёС‚Р°РµРј РґРѕР»СЋ РѕС‚ РёСЃС…РѕРґРЅРѕРіРѕ q_orig,
            # С‡С‚РѕР±С‹ РїРѕС‚РѕРјРєРё СѓС‡РёС‚С‹РІР°Р»Рё, С‡С‚Рѕ СЂРѕРґРёС‚РµР»СЏ С‡Р°СЃС‚РёС‡РЅРѕ Р·Р°РєСЂС‹Р»Рё СЃРєР»Р°РґРѕРј/РїРѕР»СѓС‡РµРЅРёСЏРјРё.
            base_qty = q_orig if np.isfinite(q_orig) else 0.0
            frac = produced / base_qty if base_qty > 1e-9 else 0.0
            frac = max(0.0, min(1.0, frac))
            scale_product[base_oid][item] = frac

            if produced <= 1e-9:
                continue  # fully covered by stock
            q_prod = int(round(produced))

            if q_prod > 0:
                key = (
                    base_oid,
                    str(r.order_id),
                    item,
                    r.due_date,
                    r.priority,
                    getattr(r, "role", "FG"),
                    getattr(r, "depth", 0),
                    wk,
                    str(getattr(r, "customer", "") or ""),
                )
                if key not in adjusted_map:
                    adjusted_map[key] = dict(
                        base_order_id=base_oid,
                        order_id=str(r.order_id),
                        item_id=item,
                        due_date=r.due_date,
                        qty=0,
                        priority=r.priority,
                        role=getattr(r, "role", "FG"),
                        depth=getattr(r, "depth", 0),
                        workshop=wk,
                        customer=str(getattr(r, "customer", "") or ""),
                    )
                adjusted_map[key]["qty"] += q_prod

        adjusted = list(adjusted_map.values())
        demand = pd.DataFrame(adjusted) if adjusted else demand.iloc[0:0]

    if len(demand) == 0:
        raise ValueError("Greedy: expanded_demand is empty вЂ” РЅРµС‚ РјР°СЂС€СЂСѓС‚РѕРІ/BOM РёР»Рё qty_per_parent/РёРґРµРЅС‚РёС„РёРєР°С‚РѕСЂС‹ РЅРµ СЃРѕС€Р»РёСЃСЊ.")
    if "role" not in demand.columns:
        demand["role"] = "FG"

    # РљР°СЂС‚С‹ РїСЂРµРґРєРѕРІ/РїРѕС‚РѕРјРєРѕРІ РґР»СЏ РґРµСЂРµРІР°
    parent_map = {}
    children_map: dict[str, set[str]] = {}
    if "root_item_id" in bom.columns:
        for r in bom[["item_id", "root_item_id"]].drop_duplicates().itertuples(index=False):
            item = str(r.item_id)
            par  = str(r.root_item_id) if pd.notna(r.root_item_id) else ""
            parent_map[item] = par or ""
            if par and par != item:
                children_map.setdefault(par, set()).add(item)
    else:
        # Р±РµР· РёРµСЂР°СЂС…РёРё вЂ” РІСЃРµ Р±РµР· СЂРѕРґРёС‚РµР»РµР№
        parent_map = {}

    # === 1) Р•РјРєРѕСЃС‚СЊ РјР°С€РёРЅ Рё РїРµСЂРµРіСЂСѓР·РєР° ===
    base_cap = machines.groupby("machine_id", as_index=True)["capacity_per_day"].max().to_dict()
    per_machine_over = {}
    if "overload_pct" in machines.columns:
        per_machine_over = machines.groupby("machine_id", as_index=True)["overload_pct"].max().to_dict()

    overrides = defaultdict(dict)  # machine_id -> {date -> cap}
    if "calendar_date" in machines.columns and "capacity_override" in machines.columns:
        mcal = machines.dropna(subset=["calendar_date", "capacity_override"])
        for _, r in mcal.iterrows():
            overrides[str(r["machine_id"])][r["calendar_date"]] = float(r["capacity_override"])

    def effective_cap(machine_id: str, day: dt.date) -> float:
        cap = overrides.get(machine_id, {}).get(day, base_cap.get(machine_id, 0.0))
        ov = per_machine_over.get(machine_id, overload_pct)
        try:
            ov = float(ov)
        except Exception:
            ov = 0.0
        ov = max(0.0, ov)
        return float(cap) * (1.0 + ov)

    # === 2) РњР°СЂС€СЂСѓС‚С‹: item_id -> [(step, machine_id, t_per_unit, setup_once)], РѕС‚СЃРѕСЂС‚РёСЂРѕРІР°РЅС‹ РїРѕ step ===
    route = defaultdict(list)
    has_setup = "setup_minutes" in bom.columns
    for _, r in bom.iterrows():
        route[str(r["item_id"])].append((
            int(r["step"]),
            str(r["machine_id"]),
            float(r["time_per_unit"]),
            float(r["setup_minutes"]) if has_setup else 0.0
        ))
    for k in route.keys():
        route[k].sort(key=lambda x: x[0])

    # РџСЂРѕРІРµСЂРёРј СЃРѕРІРїР°РґРµРЅРёРµ machine_id РјРµР¶РґСѓ BOM Рё machines
    route_mids = sorted({m for steps in route.values() for (_, m, _, _) in steps})
    base_cap = machines.groupby("machine_id")["capacity_per_day"].max()
    missing = [m for m in route_mids if m not in base_cap.index]
    print(f"[GREEDY DEBUG] route_machines={len(route_mids)} missing_in_machines={len(missing)}")
    if missing[:10]: print("[GREEDY DEBUG] missing sample:", missing[:10])
    
    # РџСЂРѕРІРµСЂРёРј РєР°Рї РЅР° СЃРµРіРѕРґРЅСЏ РїРѕ РїРµСЂРІС‹Рј 5 РјР°С€РёРЅР°Рј РёР· РјР°СЂС€СЂСѓС‚РѕРІ
    sd = start_date
    for mid in route_mids[:5]:
        try:
            c = effective_cap(mid, sd)
        except Exception:
            c = None
        print(f"[GREEDY DEBUG] cap[{mid}] on {sd} = {c}")


    # === 3) РђРєРєСѓРјСѓР»СЏС‚РѕСЂС‹ ===
    used = defaultdict(lambda: defaultdict(float))  # machine_id -> date -> minutes_used
    rows = []

    # Р’ JIT СЂРµР¶РёРјРµ РЅР°Рј РІР°Р¶РЅРѕ Р·РЅР°С‚СЊ СЃС‚Р°СЂС‚/С„РёРЅРёС€ РєР°Р¶РґРѕРіРѕ (base_order_id, item_id)
    item_start: dict[tuple[str, str], dt.date] = {}
    item_finish: dict[tuple[str, str], dt.date] = {}

    # РЈС‚РёР»РёС‚С‹ СЂР°Р·РјРµС‰РµРЅРёСЏ РѕРґРЅРѕРіРѕ РЁРђР“Рђ
    def alloc_forward_step(machine_id: str, minutes_total: float, day_from: dt.date, latest_allowed: dt.date | None) -> tuple[dt.date, dt.date]:
        """Р Р°Р·РјРµС‰Р°РµС‚ minutes_total РІРїРµСЂС‘Рґ РѕС‚ day_from. Р•СЃР»Рё latest_allowed Р·Р°РґР°РЅ, РЅРµ РґРѕРїСѓСЃРєР°РµРј day > latest_allowed."""
        remaining = float(minutes_total)
        day = max(day_from, start_date)  # РЅРµ СЂР°РЅСЊС€Рµ СЃРµРіРѕРґРЅСЏС€РЅРµРіРѕ РґРЅСЏ
        first_day = None
        guard = 0
        while remaining > 1e-6:
            guard += 1
            if guard > guard_limit_days:
                raise RuntimeError("РџСЂРµРІС‹С€РµРЅ Р»РёРјРёС‚ РґРЅРµР№ РїСЂРё РїР»Р°РЅРёСЂРѕРІР°РЅРёРё РІРїРµСЂС‘Рґ (guard_limit_days).")
            if latest_allowed is not None and day > latest_allowed:
                raise RuntimeError("РќРµ СѓРґР°Р»РѕСЃСЊ СѓР»РѕР¶РёС‚СЊСЃСЏ РІ Р·Р°РґР°РЅРЅС‹Р№ РґРµРґР»Р°Р№РЅ РїСЂРё РїР»Р°РЅРёСЂРѕРІР°РЅРёРё РІРїРµСЂС‘Рґ.")
            cap = effective_cap(machine_id, day)
            free = max(0.0, cap - used[machine_id][day])
            if free > 1e-6:
                take = min(free, remaining)
                used[machine_id][day] += take
                remaining -= take
                if first_day is None:
                    first_day = day
            if remaining > 1e-6:
                day = day + dt.timedelta(days=1)
        return first_day or day_from, day  # start_day, finish_day (РїРѕСЃР»РµРґРЅРёР№ РґРµРЅСЊ, РіРґРµ РґРѕР±РёР»Рё)

    def alloc_backward_step(machine_id: str, minutes_total: float, deadline: dt.date, earliest_allowed: dt.date) -> tuple[dt.date, dt.date]:
        """Р Р°Р·РјРµС‰Р°РµС‚ minutes_total РќРђР—РђР”, Р·Р°РєР°РЅС‡РёРІР°СЏ РІ deadline (<=deadline), РЅРѕ РЅРµ СЂР°РЅСЊС€Рµ earliest_allowed."""
        remaining = float(minutes_total)
        day = deadline
        last_day = None
        guard = 0
        while remaining > 1e-6:
            guard += 1
            if guard > guard_limit_days:
                raise RuntimeError("РџСЂРµРІС‹С€РµРЅ Р»РёРјРёС‚ РґРЅРµР№ РїСЂРё РїР»Р°РЅРёСЂРѕРІР°РЅРёРё РІРїРµСЂС‘Рґ (guard_limit_days).")
            if day < earliest_allowed:
                raise RuntimeError("РќРµ СѓРґР°Р»РѕСЃСЊ СѓР»РѕР¶РёС‚СЊСЃСЏ РІ РѕРєРЅРѕ РїСЂРё РїР»Р°РЅРёСЂРѕРІР°РЅРёРё РЅР°Р·Р°Рґ.")
            cap = effective_cap(machine_id, day)
            free = max(0.0, cap - used[machine_id][day])
            if free > 1e-6:
                take = min(free, remaining)
                used[machine_id][day] += take
                remaining -= take
                if last_day is None:
                    last_day = day
            if remaining > 1e-6:
                day = day - dt.timedelta(days=1)
        # РІРµСЂРЅС‘Рј (start_day, finish_day) РґР»СЏ С€Р°РіР°
        # start_day = РґРµРЅСЊ РЅР°С‡Р°Р»Р° (СЃР°РјС‹Р№ СЂР°РЅРЅРёР№ Р·Р°РґРµР№СЃС‚РІРѕРІР°РЅРЅС‹Р№), СЌС‚Рѕ day+1 РїРѕСЃР»Рµ РїРѕСЃР»РµРґРЅРµРіРѕ С€Р°РіР° С†РёРєР»Р°,
        # РЅРѕ РїСЂРѕС‰Рµ РѕС‚СЃР»РµРґРёС‚СЊ: РїРѕСЃР»Рµ С†РёРєР»Р° 'day' СѓР¶Рµ РЅР° 1 СЂР°РЅСЊС€Рµ С„Р°РєС‚РёС‡РµСЃРєРѕРіРѕ СЃС‚Р°СЂС‚РѕРІРѕРіРѕ РґРЅСЏ
        start_day = day + dt.timedelta(days=1)
        finish_day = last_day or deadline
        return start_day, finish_day

    # Р Р°Р·РјРµС‰РµРЅРёРµ РѕРґРЅРѕРіРѕ РР—Р”Р•Р›РРЇ РїРѕ РІСЃРµРј РµРіРѕ С€Р°РіР°Рј
    def schedule_item_forward(base_oid: str, oid: str, item: str, qty: int, earliest: dt.date, latest: dt.date | None, due: dt.date):
        """ASAP: С€Р°РіР°РµРј РІРїРµСЂС‘Рґ, optionally СѓРґРµСЂР¶РёРІР°СЏ finish <= latest."""
        steps = route.get(item) or [(1, "UNKNOWN", 0.0, 0.0)]
        first_any = None
        cur_day = earliest
        for step, machine_id, tpu, setup_once in steps:
            total = qty * float(tpu) + float(setup_once)
            st, fin = alloc_forward_step(machine_id, total, cur_day, latest_allowed=latest)
            # Р—Р°РїРёС€РµРј РїРѕ РґРЅСЏРј (С‚Р°Рј СѓР¶Рµ РґРѕР±Р°РІРёР»РёСЃСЊ minutes РІ used) вЂ” РґРѕР±Р°РІРёРј СЃС‚СЂРѕРєРё
            # Р§С‚РѕР±С‹ РЅРµ СЂР°Р·Р±РёСЂР°С‚СЊ РїРѕ РјРёРЅСѓС‚Р°Рј РІ РїСЂРµРґРµР»Р°С… РґРЅСЏ, РјС‹ СѓР¶Рµ РїРёС€РµРј РїРѕ С„Р°РєС‚Сѓ РІ while-С†РёРєР»Рµ РІС‹С€Рµ;
            # Р·РґРµСЃСЊ С„РёРєСЃРёСЂСѓРµРј С‚РѕР»СЊРєРѕ РіСЂР°РЅРёС†С‹ РґР»СЏ Р·Р°РІРёСЃРёРјРѕСЃС‚РµР№:
            cur_day = fin  # СЃР»РµРґСѓСЋС‰РёР№ С€Р°Рі РЅРµ СЂР°РЅСЊС€Рµ С„РёРЅРёС€Р° С‚РµРєСѓС‰РµРіРѕ
            first_any = first_any or st
            # (РЎС‚СЂРѕРєРё СЂР°СЃРїРёСЃР°РЅРёСЏ СѓР¶Рµ СЃРѕР±РёСЂР°Р»РёСЃСЊ РІ alloc_* С‡РµСЂРµР· used; РґРѕР±Р°РІРёРј РёС… Р·РґРµСЃСЊ Р°РіСЂРµРіРёСЂРѕРІР°РЅРЅРѕ РїРѕ РґРЅСЏРј)
        # Р”Р»СЏ РєРѕСЂСЂРµРєС‚РЅРѕРіРѕ СЌРєСЃРїРѕСЂС‚Р° Рё lag вЂ” РЅР°Рј РЅСѓР¶РЅС‹ СЂРµР°Р»СЊРЅС‹Рµ СЃС‚СЂРѕРєРё РїРѕ РґРЅСЏРј.
        # РњС‹ СѓР¶Рµ РёСЃРїРѕР»СЊР·РѕРІР°Р»Рё used[...] РґР»СЏ СЂРµР·РµСЂРІР°, РЅРѕ СЃС‚СЂРѕС‡РµРє РЅРµ РґРѕР±Р°РІРёР»Рё. Р”РѕР±Р°РІРёРј РёС… СЃРµР№С‡Р°СЃ РїРѕСЃС‚С„Р°РєС‚СѓРј:
        # (РїСЂРѕР№РґС‘Рј РїРѕ РґРёР°РїР°Р·РѕРЅСѓ РґР°С‚ Рё РІС‹С‡Р»РµРЅРёРј РІРєР»Р°Рґ СЌС‚РѕРіРѕ item/oid вЂ” СѓРїСЂРѕСЃС‚РёРј: РґРѕР±Р°РІРёРј СЃС‚СЂРѕРєРё РІРѕ РІСЂРµРјСЏ СЂР°СЃРїСЂРµРґРµР»РµРЅРёСЏ)
        # => РџРѕСЌС‚РѕРјСѓ РїРµСЂРµРЅРѕСЃРёРј С„РѕСЂРјРёСЂРѕРІР°РЅРёРµ СЃС‚СЂРѕРє РІРЅСѓС‚СЂСЊ alloc_* (СЃРј. РЅРёР¶Рµ РѕР±РЅРѕРІР»С‘РЅРЅС‹Р№ РІР°СЂРёР°РЅС‚).
        item_start[(base_oid, item)] = first_any or earliest
        item_finish[(base_oid, item)] = cur_day
        return item_start[(base_oid, item)], item_finish[(base_oid, item)]

    # Р§С‚РѕР±С‹ С„РѕСЂРјРёСЂРѕРІР°С‚СЊ СЃС‚СЂРѕРєРё СЂР°СЃРїРёСЃР°РЅРёСЏ СЃСЂР°Р·Сѓ, СЃРґРµР»Р°РµРј РїСЂРѕРєСЃРё РІРѕРєСЂСѓРі alloc_*,
    # РєРѕС‚РѕСЂС‹Р№ РїСЂРёРЅРёРјР°РµС‚ "РєРѕРЅС‚РµРєСЃС‚ СЃС‚СЂРѕРєРё" Рё РїСЂРё РєР°Р¶РґРѕРј 'take' РґРѕР±Р°РІР»СЏРµС‚ row.
    # ---------- РђР»Р»РѕРєР°С‚РѕСЂС‹ (РІРїРµСЂС‘Рґ / РЅР°Р·Р°Рґ) ----------
    def alloc_forward_step_rows(machine_id, minutes_total, day_from, latest_allowed, row_ctx):
        # robust numeric
        try:
            remaining = float(minutes_total)
        except Exception:
            remaining = 0.0
        if not np.isfinite(remaining) or remaining < 0:
            remaining = 0.0

        day = max(day_from, start_date)
        first_day = None
        guard = 0

        while remaining > 1e-6:
            guard += 1
            if guard > guard_limit_days:
                raise RuntimeError("РџСЂРµРІС‹С€РµРЅ Р»РёРјРёС‚ РґРЅРµР№ РїСЂРё РїР»Р°РЅРёСЂРѕРІР°РЅРёРё РІРїРµСЂС‘Рґ (guard_limit_days).")
            if latest_allowed is not None and day > latest_allowed:
                raise RuntimeError("РќРµ СѓРґР°Р»РѕСЃСЊ СѓР»РѕР¶РёС‚СЊСЃСЏ РІ РґРµРґР»Р°Р№РЅ РїСЂРё РїР»Р°РЅРёСЂРѕРІР°РЅРёРё РІРїРµСЂС‘Рґ.")

            cap = effective_cap(machine_id, day)
            free = max(0.0, cap - used[machine_id][day])

            if free > 1e-6:
                take = min(free, remaining)
                used[machine_id][day] += take
                rows.append({**row_ctx, "machine_id": machine_id, "date": day, "minutes": take})
                remaining -= take
                if first_day is None:
                    first_day = day

            if remaining > 1e-6:
                day = day + dt.timedelta(days=1)

        return first_day or day_from, day


    def alloc_backward_step_rows(machine_id, minutes_total, deadline, earliest_allowed, row_ctx):
        # robust numeric
        try:
            remaining = float(minutes_total)
        except Exception:
            remaining = 0.0
        if not np.isfinite(remaining) or remaining < 0:
            remaining = 0.0
        if remaining <= 1e-6:
            return deadline, deadline

        # Validate feasibility in [earliest_allowed, deadline] before mutating `used`.
        if deadline < earliest_allowed:
            raise RuntimeError("Не удалось уложиться в окно при планировании назад.")
        span_days = (deadline - earliest_allowed).days + 1
        if span_days > guard_limit_days:
            raise RuntimeError("Превышен лимит дней при планировании назад (guard_limit_days).")
        free_total = 0.0
        probe_day = deadline
        while probe_day >= earliest_allowed:
            cap = effective_cap(machine_id, probe_day)
            free_total += max(0.0, cap - used[machine_id][probe_day])
            probe_day = probe_day - dt.timedelta(days=1)
        if free_total + 1e-6 < remaining:
            raise RuntimeError("Не удалось уложиться в окно при планировании назад.")

        day = deadline
        last_day = None
        guard = 0

        while remaining > 1e-6:
            guard += 1
            if guard > guard_limit_days:
                raise RuntimeError("РџСЂРµРІС‹С€РµРЅ Р»РёРјРёС‚ РґРЅРµР№ РїСЂРё РїР»Р°РЅРёСЂРѕРІР°РЅРёРё РЅР°Р·Р°Рґ (guard_limit_days).")
            if day < earliest_allowed:
                raise RuntimeError("Не удалось уложиться в окно при планировании назад.")

            cap = effective_cap(machine_id, day)
            free = max(0.0, cap - used[machine_id][day])

            if free > 1e-6:
                take = min(free, remaining)
                used[machine_id][day] += take
                rows.append({**row_ctx, "machine_id": machine_id, "date": day, "minutes": take})
                remaining -= take
                if last_day is None:
                    last_day = day

            if remaining > 1e-6:
                day = day - dt.timedelta(days=1)

        start_day = day
        finish_day = last_day or deadline
        return start_day, finish_day


    def schedule_item_backward(base_oid: str, oid: str, item: str, qty: int,
                               deadline: dt.date, earliest_allowed: dt.date, due: dt.date):
        """JIT-С…РµР»РїРµСЂ: РѕРґРёРЅ item РЅР°Р·Р°Рґ Рє deadline (<=deadline), Р±РµР· СЃС‚Р°СЂС‚Р° СЂР°РЅСЊС€Рµ earliest_allowed."""
        steps = route.get(item) or [(1, "UNKNOWN", 0.0, 0.0)]
        cur_deadline = deadline
        earliest_seen = None

        # СЂРѕРІРЅРѕ Рё Р±РµР·РѕРїР°СЃРЅРѕ СЃС‡РёС‚Р°РµРј РјРёРЅСѓС‚С‹ РїРѕ С€Р°РіР°Рј
        def _num(x): 
            try:
                x = float(x)
            except Exception:
                x = 0.0
            if not np.isfinite(x): x = 0.0
            return x

        for step, machine_id, tpu, setup_once in reversed(steps):
            tpu = _num(tpu); setup_once = _num(setup_once); q = int(qty) if not pd.isna(qty) else 0
            total = q * tpu + setup_once

            row_ctx = {
                "base_order_id": base_oid,
                "order_id": oid,
                "item_id": item,
                "step": step,
                "qty": q,
                "due_date": due, "workshop": item_workshop.get(item, "")}
            st, fin = alloc_backward_step_rows(machine_id, total, cur_deadline, earliest_allowed, row_ctx)
            earliest_seen = st if earliest_seen is None else min(earliest_seen, st)
            # РїСЂРµРґС‹РґСѓС‰РёР№ РїРѕ РјР°СЂС€СЂСѓС‚Сѓ С€Р°Рі РґРѕР»Р¶РµРЅ С„РёРЅРёС€РёСЂРѕРІР°С‚СЊ РЅРµ РїРѕР·Р¶Рµ СЃС‚Р°СЂС‚Р° С‚РµРєСѓС‰РµРіРѕ - 1 РґРµРЅСЊ
            cur_deadline = st - dt.timedelta(days=1)

        item_start[(base_oid, item)] = earliest_seen or earliest_allowed
        item_finish[(base_oid, item)] = deadline
        return item_start[(base_oid, item)], item_finish[(base_oid, item)]


    # === 4) РћСЃРЅРѕРІРЅР°СЏ Р»РѕРіРёРєР°: ASAP РёР»Рё JIT ===
    if not align_roots_to_due:
        # ---------- ASAP РІРїРµСЂС‘Рґ ----------
        sort_cols = [c for c in ["priority", "base_order_id", "item_id"] if c in demand.columns]
        demand_sorted = demand.sort_values(sort_cols, kind="stable").reset_index(drop=True)

        for _, job in demand_sorted.iterrows():
            base_oid = str(job.get("base_order_id", job["order_id"]))
            oid = str(job["order_id"])
            item = str(job["item_id"])
            qty = int(job["qty"]) if not pd.isna(job["qty"]) else 0
            due = job["due_date"]

            # СЃС‚Р°СЂС‚ РЅРµ СЂР°РЅСЊС€Рµ С„РёРЅРёС€Р° СЂРѕРґРёС‚РµР»СЏ
            earliest_day = start_date
            par = str(parent_map.get(item, "") or "")
            if par:
                pf = item_finish.get((base_oid, par))
                if pf:
                    earliest_day = max(earliest_day, pf)

            steps = route.get(item) or [(1, "UNKNOWN", 0.0, 0.0)]

            # Р°РєРєСѓСЂР°С‚РЅРѕ СЃС‡РёС‚Р°РµРј СЃСѓРјРјР°СЂРЅС‹Рµ РјРёРЅСѓС‚С‹
            def _num(x):
                try:
                    x = float(x)
                except Exception:
                    x = 0.0
                if not np.isfinite(x): x = 0.0
                return x
            total_item_minutes = 0.0
            for _, _, tpu, setup in steps:
                total_item_minutes += qty * _num(tpu) + _num(setup)

            if total_item_minutes <= 1e-9:
                # РџСѓСЃС‚РѕР№ РјР°СЂС€СЂСѓС‚ вЂ” РїР»РµР№СЃС…РѕР»РґРµСЂ
                rows.append({
                    "base_order_id": base_oid, "order_id": oid, "item_id": item,
                    "step": 1, "machine_id": "UNKNOWN", "date": earliest_day,
                    "minutes": 0.0, "qty": qty, "due_date": due,
                })
                item_start[(base_oid, item)] = earliest_day
                item_finish[(base_oid, item)] = earliest_day
                warnings.append(f"[ROUTE] {base_oid}/{item}: РїСѓСЃС‚РѕР№ РјР°СЂС€СЂСѓС‚ (0 РјРёРЅ). РџР»РµР№СЃС…РѕР»РґРµСЂ {earliest_day}.")
                continue

            # СЃРѕР±СЃС‚РІРµРЅРЅРѕ РїР»Р°РЅРёСЂРѕРІР°РЅРёРµ РІРїРµСЂС‘Рґ
            first_any = None
            cur = earliest_day  # <--- РїСЂРѕРїСѓС‰РµРЅРЅР°СЏ РїРµСЂРµРјРµРЅРЅР°СЏ Р±С‹Р»Р°
            for step, machine_id, tpu, setup_once in steps:
                tpu = _num(tpu); setup_once = _num(setup_once)
                total = qty * tpu + setup_once
                row_ctx = {
                    "base_order_id": base_oid, "order_id": oid, "item_id": item,
                    "step": step, "qty": qty, "due_date": due,
                }
                st, fin = alloc_forward_step_rows(machine_id, total, cur, latest_allowed=None, row_ctx=row_ctx)
                cur = fin
                if first_any is None:
                    first_any = st

            item_start[(base_oid, item)] = first_any or earliest_day
            item_finish[(base_oid, item)] = cur

    else:
        # ---------- JIT: РєРѕСЂРЅРµРІС‹Рµ (FG) СѓРїРёСЂР°СЋС‚СЃСЏ РІ due_date, РґРµС‚Рё РЅР°Р·Р°Рґ РґРѕ СЃС‚Р°СЂС‚Р° СЂРѕРґРёС‚РµР»СЏ ----------
        today = start_date  # РЅРµ СЃС‚Р°СЂС‚СѓРµРј СЂР°РЅСЊС€Рµ today

        # РіСЂСѓРїРїРёСЂСѓРµРј РїРѕ base_order_id (РµСЃР»Рё РµРіРѕ РЅРµС‚ вЂ” РїРѕ order_id)
        grp_cols = ["base_order_id"] if "base_order_id" in demand.columns else ["order_id"]
        for base_oid, df_ord in demand.groupby(grp_cols):
            base_oid = str(base_oid if isinstance(base_oid, str) else base_oid[0])

            # РєРѕСЂРЅРё РїРѕ role='FG', РёРЅР°С‡Рµ вЂ” Сѓ РєРѕРіРѕ РЅРµС‚ СЂРѕРґРёС‚РµР»СЏ
            roots = df_ord[df_ord["role"] == "FG"]["item_id"].astype(str).unique().tolist()
            if not roots:
                roots = [it for it in df_ord["item_id"].astype(str).unique().tolist() if not parent_map.get(it, "")]
                if not roots and len(df_ord):
                    roots = [str(df_ord["item_id"].astype(str).iloc[0])]

            # РїР»Р°РЅРёСЂРѕРІР°С‚СЊ РєР°Р¶РґС‹Р№ РєРѕСЂРµРЅСЊ РїРѕ РљРђР–Р”РћРњРЈ РµРіРѕ due РІ СЌС‚РѕР№ РіСЂСѓРїРїРµ
            for root in roots:
                # РІСЃРµ due РґР»СЏ root РІ СЌС‚РѕР№ РіСЂСѓРїРїРµ
                cand_root = df_ord[df_ord["item_id"].astype(str) == root].copy()
                if cand_root.empty:
                    continue
                cand_root["due_date"] = pd.to_datetime(cand_root["due_date"])
                for due_ts in sorted(cand_root["due_date"].unique()):
                    deadline = due_ts.date()

                    def plan_down(item_id: str, deadline: dt.date):
                        # РІС‹Р±СЂР°С‚СЊ СЃС‚СЂРѕРєРё СЃРїСЂРѕСЃР° РґР»СЏ item_id (РІСЃРµ, РїРѕ Р±Р»РёР¶Р°Р№С€РµР№ Рє РґРµРґР»Р°Р№РЅСѓ)
                        cc = df_ord[df_ord["item_id"].astype(str) == item_id].copy()
                        if cc.empty:
                            return
                        cc["due_date"] = pd.to_datetime(cc["due_date"])
                        target = pd.to_datetime(deadline)
                        cc = cc.assign(_dist=(cc["due_date"] - target).abs()).sort_values(
                            ["_dist", "due_date", "order_id"], kind="stable"
                        ).drop(columns=["_dist"])

                        # СЃСѓРјРјР°СЂРЅС‹Рµ РјРёРЅСѓС‚С‹ (robust)
                        def _num(x):
                            try:
                                x = float(x)
                            except Exception:
                                x = 0.0
                            if not np.isfinite(x): x = 0.0
                            return x

                        min_start = None
                        max_finish = None
                        for _, row in cc.iterrows():
                            q = int(row["qty"]) if not pd.isna(row["qty"]) else 0
                            due_loc = row["due_date"].date()
                            oid_loc = str(row["order_id"])

                            steps_loc = route.get(item_id) or [(1, "UNKNOWN", 0.0, 0.0)]
                            total_item_minutes = 0.0
                            for _, _, tpu_i, setup_i in steps_loc:
                                total_item_minutes += q * _num(tpu_i) + _num(setup_i)

                            if total_item_minutes <= 1e-9:
                                place_day = deadline if deadline >= today else today
                                rows.append({
                                    "base_order_id": base_oid, "order_id": oid_loc, "item_id": item_id,
                                    "step": 1, "machine_id": "UNKNOWN", "date": place_day,
                                    "minutes": 0.0, "qty": q, "due_date": due_loc,
                                })
                                start_fact = place_day
                                finish_fact = place_day
                            else:
                                try:
                                    earliest_seen = None
                                    cur_deadline = deadline

                                    for step, machine_id, tpu, setup in reversed(steps_loc):
                                        tpu = _num(tpu); setup = _num(setup)
                                        total = q * tpu + setup
                                        ctx = {"base_order_id": base_oid, "order_id": oid_loc,
                                               "item_id": item_id, "step": step, "qty": q, "due_date": due_loc, "workshop": item_workshop.get(item_id, "")}
                                        st, fin = alloc_backward_step_rows(machine_id, total, cur_deadline, earliest_allowed=today, row_ctx=ctx)
                                        earliest_seen = st if earliest_seen is None else min(earliest_seen, st)
                                        cur_deadline = st - dt.timedelta(days=1)

                                    start_fact = earliest_seen or today
                                    finish_fact = deadline
                                except RuntimeError:
                                    # If JIT window [today, deadline] is infeasible, place overdue item forward from today.
                                    first_any = None
                                    cur = today
                                    for step, machine_id, tpu, setup in steps_loc:
                                        tpu = _num(tpu); setup = _num(setup)
                                        total = q * tpu + setup
                                        ctx = {"base_order_id": base_oid, "order_id": oid_loc,
                                               "item_id": item_id, "step": step, "qty": q, "due_date": due_loc, "workshop": item_workshop.get(item_id, "")}
                                        st, fin = alloc_forward_step_rows(machine_id, total, cur, latest_allowed=None, row_ctx=ctx)
                                        cur = fin
                                        if first_any is None:
                                            first_any = st
                                    start_fact = first_any or today
                                    finish_fact = cur

                            min_start = start_fact if min_start is None else min(min_start, start_fact)
                            max_finish = finish_fact if max_finish is None else max(max_finish, finish_fact)

                        if min_start is None:
                            return
                        item_start[(base_oid, item_id)] = min_start
                        item_finish[(base_oid, item_id)] = max_finish

                        child_deadline = item_start[(base_oid, item_id)] - dt.timedelta(days=1)
                        for child in sorted(list(children_map.get(item_id, []))):
                            plan_down(child, deadline=child_deadline)

                    # РїР»Р°РЅРёСЂСѓРµРј РєРѕСЂРµРЅСЊ РЅР° СЌС‚РѕС‚ РєРѕРЅРєСЂРµС‚РЅС‹Р№ due
                    plan_down(root, deadline=deadline)


    # === 5) Р¤РёРЅР°Р»РёР·Р°С†РёСЏ СЂР°СЃРїРёСЃР°РЅРёСЏ ===
    sched = pd.DataFrame(rows)
    if sched.empty:
        sched = pd.DataFrame([{
            "base_order_id": "NO_JOBS",
            "order_id": "NO_JOBS",
            "item_id": "",
            "step": 1,
            "machine_id": "UNKNOWN",
            "date": start_date,
            "minutes": 0.0,
            "qty": 0,
            "due_date": start_date,
        }])

    sched["lag_days"] = (pd.to_datetime(sched["date"]) - pd.to_datetime(sched["due_date"])).dt.days
    sched = sched.sort_values(["date","machine_id","order_id","step"], kind="stable").reset_index(drop=True)

    if not rows:
        print("[GREEDY INFO] Р’СЃРµ РїРѕС‚СЂРµР±РЅРѕСЃС‚Рё РїРѕРєСЂС‹С‚С‹ Р·Р°РїР°СЃРѕРј вЂ” РїСЂРѕРёР·РІРѕРґСЃС‚РІРµРЅРЅС‹Рµ РѕРїРµСЂР°С†РёРё РЅРµ СЃРѕР·РґР°РІР°Р»РёСЃСЊ.")
        sched = pd.DataFrame([{
            "base_order_id": "ALL_FROM_STOCK",
            "order_id": "ALL_FROM_STOCK",
            "item_id": "",
            "step": 1,
            "machine_id": "NONE",
            "date": start_date,
            "minutes": 0.0,
            "qty": 0,
            "due_date": start_date,
        }])
        sched["lag_days"] = 0
        try:
            sched.attrs["warnings"] = warnings + ["ALL_FROM_STOCK"]
        except Exception:
            pass
        return sched
    print("[GREEDY DEBUG] rows_appended:", len(rows))

    try:
        sched.attrs["warnings"] = warnings
    except Exception:
        pass

    return sched


# =========================
# Export (Excel + stacked bar)
# =========================
def _auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(60, max(12, max_len + 2))


from openpyxl.chart import BarChart, Reference
from openpyxl.chart.layout import Layout, ManualLayout

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter
import pandas as pd
from pathlib import Path

def compute_order_items_timeline(sched: pd.DataFrame) -> pd.DataFrame:
    """
    РўР°Р№РјР»Р°Р№РЅ РґР»СЏ РєР°Р¶РґРѕР№ РїР°СЂС‹ (order_id, item_id):
      start_date, finish_date, duration_days, due_date, finish_lag
    """
    if sched.empty:
        return pd.DataFrame(columns=["order_id","item_id","start_date","finish_date","duration_days","due_date","finish_lag"])

    g = sched.groupby(["order_id","item_id"], as_index=False).agg(
        start_date=("date", "min"),
        finish_date=("date", "max"),
        due_date=("due_date", "max"),
    )
    g["duration_days"] = (pd.to_datetime(g["finish_date"]) - pd.to_datetime(g["start_date"])).dt.days + 1
    g["finish_lag"] = (pd.to_datetime(g["finish_date"]) - pd.to_datetime(g["due_date"])).dt.days
    return g.sort_values(["order_id","item_id","start_date"], kind="stable").reset_index(drop=True)


def export_with_charts(sched: pd.DataFrame, out_xlsx: Path, bom: pd.DataFrame | None = None) -> Path:
    sched = sched.copy()
    sched = sched.copy()
    for c in ("minutes", "qty"):
        if c in sched.columns:
            sched[c] = pd.to_numeric(sched[c], errors="coerce").fillna(0.0)
    
    by_day = sched.groupby(["date", "machine_id"], as_index=False)["minutes"].sum()
    by_day["minutes"] = pd.to_numeric(by_day["minutes"], errors="coerce").fillna(0.0)
    gmat = by_day.pivot(index="machine_id", columns="date", values="minutes").fillna(0).sort_index(axis=0)

    wb = Workbook()
    ws_sched = wb.active; ws_sched.title = "schedule"
    cols = ["base_order_id","order_id","item_id","workshop","step","machine_id","date","minutes","qty","due_date","lag_days"]
    cols = [c for c in cols if c in sched.columns]
    ws_sched.append(cols)
    for _, r in sched[cols].iterrows():
        ws_sched.append(list(r.values))
    _auto_width(ws_sched)

    ws_alloc = wb.create_sheet("alloc_by_day")
    ws_alloc.append(["date","machine_id","minutes"])
    for _, r in by_day.iterrows():
        ws_alloc.append([r["date"], r["machine_id"], float(r["minutes"])])
    _auto_width(ws_alloc)

    ws_gmat = wb.create_sheet("gantt_matrix")
    ws_gmat.append(["machine_id"] + [d for d in gmat.columns.tolist()])
    for mid, row in gmat.iterrows():
        ws_gmat.append([mid] + [float(x) for x in row.tolist()])
    _auto_width(ws_gmat)
    if gmat.shape[0] > 0 and gmat.shape[1] > 0:
        chart = BarChart(); chart.type="col"; chart.grouping="stacked"
        chart.title = "Р—Р°РіСЂСѓР·РєР° (РјРёРЅСѓС‚С‹ РїРѕ РјР°С€РёРЅР°Рј/РґРЅСЏРј)"; chart.y_axis.title="РњРёРЅСѓС‚С‹"; chart.x_axis.title="Р”Р°С‚Р°"
        data_ref = Reference(ws_gmat, min_col=2, min_row=1, max_col=1+gmat.shape[1], max_row=1+gmat.shape[0])
        cats_ref = Reference(ws_gmat, min_col=2, min_row=1, max_col=1+gmat.shape[1], max_row=1)
        chart.add_data(data_ref, titles_from_data=True, from_rows=True); chart.set_categories(cats_ref)
        chart.height=18; chart.width=28
        ws_gmat.add_chart(chart, f"B{gmat.shape[0] + 4}")

    # --- util_heatmap_% ---
    ws_util = wb.create_sheet("util_heatmap_%")
    try:
        cap_tbl = _build_cap_matrix(export_with_charts._machines_df) if hasattr(export_with_charts, "_machines_df") else None
    except Exception:
        cap_tbl = None
    if cap_tbl is None or cap_tbl.empty:
        cap_tbl = pd.DataFrame({"machine_id": by_day["machine_id"].unique(), "cap_minutes": 0.0})
    
    if "calendar_date" in cap_tbl.columns:
        cap_by_date = cap_tbl.rename(columns={"calendar_date":"date"})
        cap_join = by_day.merge(cap_by_date[["machine_id","date","cap_minutes"]],
                                on=["machine_id","date"], how="left")
        base_cap = cap_tbl.groupby("machine_id", as_index=False)["cap_minutes"].max() \
                          .rename(columns={"cap_minutes":"base_cap"})
        cap_join = cap_join.merge(base_cap, on="machine_id", how="left")
        cap_join["cap_eff"] = cap_join["cap_minutes"].fillna(cap_join["base_cap"]).fillna(0.0)
    else:
        base_cap = cap_tbl.groupby("machine_id", as_index=False)["cap_minutes"].max() \
                          .rename(columns={"cap_minutes":"cap_eff"})
        cap_join = by_day.merge(base_cap, on="machine_id", how="left")
        cap_join["cap_eff"] = cap_join["cap_eff"].fillna(0.0)
    
    # <-- Р’РђР–РќРћ: РїСЂРёРІРµРґРµРЅРёРµ Рє С‡РёСЃР»Р°Рј Р”O Р›Р®Р‘РћР™ РђР РР¤РњР•РўРРљР
    cap_join["minutes"] = pd.to_numeric(cap_join["minutes"], errors="coerce").fillna(0.0)
    cap_join["cap_eff"] = pd.to_numeric(cap_join["cap_eff"], errors="coerce").fillna(0.0)
    
    # РЎС‡РёС‚Р°РµРј % Р·Р°РіСЂСѓР·РєРё РѕРґРёРЅ СЂР°Р·, Р±РµР·РѕРїР°СЃРЅРѕ
    cap_join["util_pct"] = np.where(
        cap_join["cap_eff"] > 0,
        (cap_join["minutes"] / cap_join["cap_eff"] * 100).round(1),
        0.0
    )
    
    util_piv = cap_join.pivot(index="machine_id", columns="date", values="util_pct").fillna(0)

    ws_util.append(["machine_id"] + [d for d in util_piv.columns.tolist()])
    for mid, row in util_piv.iterrows():
        ws_util.append([mid] + [float(x) for x in row.tolist()])
    _auto_width(ws_util)
    if util_piv.shape[0] > 0 and util_piv.shape[1] > 0:
        from openpyxl.utils import get_column_letter as _gcl
        min_row = 2; max_row = 1 + util_piv.shape[0]
        min_col = 2; max_col = 1 + util_piv.shape[1]
        cell_range = f"{_gcl(min_col)}{min_row}:{_gcl(max_col)}{max_row}"
        rule = ColorScaleRule(start_type='num', start_value=0, start_color='FFFFFF',
                              mid_type='num', mid_value=50, mid_color='FFF59D',
                              end_type='num', end_value=100, end_color='F44336')
        ws_util.conditional_formatting.add(cell_range, rule)
        ws_util.freeze_panes = "B2"

    # Orders rollup (per order_id/item_id/workshop)
    ws_orders = wb.create_sheet("Orders")
    ws_orders.append(["order_id", "item_id", "workshop", "qty", "date_start", "date_finish"])
    try:
        df_orders = sched.copy()
        if "order_id" not in df_orders.columns or "item_id" not in df_orders.columns or "date" not in df_orders.columns:
            raise ValueError("missing core columns")
        if "workshop" not in df_orders.columns:
            df_orders["workshop"] = ""
        df_orders = df_orders.dropna(subset=["order_id", "item_id"])
        df_orders["date"] = pd.to_datetime(df_orders["date"], errors="coerce")
        df_orders = df_orders.dropna(subset=["date"])
        df_orders["qty"] = pd.to_numeric(df_orders.get("qty", 0), errors="coerce").fillna(0.0)
        # qty of the order (not sum of all ops): take max to avoid double counting across steps/days
        grouped = df_orders.groupby(["order_id", "item_id", "workshop"], as_index=False).agg(
            qty=("qty", "max"),
            date_start=("date", "min"),
            date_finish=("date", "max"),
        )
        grouped["date_start"] = pd.to_datetime(grouped["date_start"]).dt.date
        grouped["date_finish"] = pd.to_datetime(grouped["date_finish"]).dt.date
        grouped = grouped.sort_values(["order_id", "item_id", "date_start"], kind="stable")
        for _, r in grouped.iterrows():
            ws_orders.append([
                r.get("order_id", ""),
                r.get("item_id", ""),
                r.get("workshop", ""),
                float(r.get("qty", 0.0)),
                r.get("date_start", ""),
                r.get("date_finish", ""),
            ])
    except Exception:
        pass
    _auto_width(ws_orders)

    orders = compute_orders_timeline(sched)
    ws_otl = wb.create_sheet("orders_timeline")
    ws_otl.append(["order_id","item_id","start_date","finish_date","duration_days","due_date","finish_lag"])
    for _, r in orders.iterrows():
        ws_otl.append(list(r.values))
    _auto_width(ws_otl)
    if not orders.empty:
        min_start = pd.to_datetime(orders["start_date"]).min().date()
        ws_otl.cell(row=1, column=8, value="offset_days")
        for i, (_, rr) in enumerate(orders.iterrows(), start=2):
            off = (pd.to_datetime(rr["start_date"]).date() - min_start).days
            ws_otl.cell(row=i, column=8, value=int(off))
        nrows = orders.shape[0]
        cats_ref = Reference(ws_otl, min_col=1, min_row=2, max_row=1 + nrows)
        data_ref1 = Reference(ws_otl, min_col=8, min_row=1, max_col=8, max_row=1 + nrows)
        data_ref2 = Reference(ws_otl, min_col=5, min_row=1, max_col=5, max_row=1 + nrows)
        gant = BarChart(); gant.type = "bar"; gant.grouping = "stacked"
        gant.title = f"Orders Gantt (t0 = {min_start})"; gant.y_axis.title = "order_id"; gant.x_axis.title = "days from t0"
        gant.add_data(data_ref1, titles_from_data=True); gant.add_data(data_ref2, titles_from_data=True); gant.set_categories(cats_ref)
        gant.height = 20; gant.width = 30
        try:
            ser_offset = gant.series[0]
            ser_offset.graphicalProperties.solidFill = "DDDDDD"
            ser_offset.graphicalProperties.line.solidFill = "DDDDDD"
        except Exception:
            pass
        ws_otl.add_chart(gant, "J2")

    oitl = compute_order_items_timeline(sched)
    ws_oitl = wb.create_sheet("order_items_timeline")
    ws_oitl.append(["order_id","item_id","start_date","finish_date","duration_days","due_date","finish_lag"])
    for _, r in oitl.iterrows():
        ws_oitl.append(list(r.values))
    _auto_width(ws_oitl)

    ws_bom = wb.create_sheet("BOM")
    if bom is not None and "root_item_id" in bom.columns:
        try:
            b = bom[["item_id","root_item_id","qty_per_parent"]].drop_duplicates()
            b["root_item_id"] = b["root_item_id"].fillna("").astype(str)
            ws_bom.append(["root_item_id","item_id","qty_per_parent"])
            for _, r in b.iterrows():
                ws_bom.append([r["root_item_id"], r["item_id"], float(r["qty_per_parent"])])
        except Exception:
            ws_bom.append(["info"]); ws_bom.append(["РќРµ СѓРґР°Р»РѕСЃСЊ СЃРѕР±СЂР°С‚СЊ BOM view."])
    else:
        ws_bom.append(["info"]); ws_bom.append(["BOM Р±РµР· РёРµСЂР°СЂС…РёРё."])
    _auto_width(ws_bom)

    if "base_order_id" in sched.columns:
        link = (sched[["base_order_id","order_id","item_id"]].drop_duplicates().sort_values(["base_order_id","order_id","item_id"], kind="stable"))
        ws_link = wb.create_sheet("orders_linkage")
        ws_link.append(["base_order_id","order_id","item_id"])
        for _, r in link.iterrows():
            ws_link.append([r["base_order_id"], r["order_id"], r["item_id"]])
        _auto_width(ws_link)

    # Child-from-FG linkage (explicit mapping FG -> CHILD with quantities)
    try:
        if set(["base_order_id","order_id","item_id"]).issubset(set(sched.columns)):
            df = sched.copy()
            df["base_order_id"] = df["base_order_id"].astype(str)
            df["order_id"] = df["order_id"].astype(str)
            df["item_id"] = df["item_id"].astype(str)
            # Identify FG rows (order_id == base_order_id)
            fg_rows = df[df["order_id"] == df["base_order_id"]]
            fg_map = fg_rows.drop_duplicates("base_order_id")[["base_order_id","item_id"]] \
                             .set_index("base_order_id")["item_id"].to_dict()
            # Child rows where order_id != base_order_id
            ch = df[df["order_id"] != df["base_order_id"]].copy()
            if not ch.empty:
                ch["fg_item"] = ch["base_order_id"].map(fg_map).fillna("")
                # customer from base prefix if present (e.g., 'Petrol-YYYYMMDD-PV...')
                def _cust_from_base(s: str) -> str:
                    try:
                        return str(s).split("-", 1)[0]
                    except Exception:
                        return ""
                ch["customer"] = ch["base_order_id"].map(_cust_from_base)
                # aggregate by base, fg_item, child, due_date
                keep_due = "due_date" if "due_date" in ch.columns else None
                agg_cols = ["base_order_id","fg_item","item_id"] + ([keep_due] if keep_due else [])
                ch_qty = ch.copy()
                ch_qty["qty"] = pd.to_numeric(ch_qty.get("qty", 0), errors="coerce").fillna(0.0)
                grp = ch_qty.groupby(agg_cols, as_index=False)["qty"].sum()
                ws_cff = wb.create_sheet("child_from_fg")
                header = ["base_order_id","customer","fg_item","child_item","due_date","qty"]
                ws_cff.append(header)
                for _, r in grp.iterrows():
                    ws_cff.append([
                        r.get("base_order_id",""),
                        ch.loc[ch["base_order_id"]==r.get("base_order_id",""), "customer"].iloc[0] if not ch.empty else "",
                        r.get("fg_item",""),
                        r.get("item_id",""),
                        (str(r.get("due_date")) if keep_due else ""),
                        float(r.get("qty", 0.0)),
                    ])
                _auto_width(ws_cff)
    except Exception:
        pass

    out_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_xlsx)
    return out_xlsx


# =========================
# Pipeline
# =========================


# =========================
# Backward-compatible wrapper (tolerates SQLAlchemy Session)
# =========================


def _parse_args():
    p = argparse.ArgumentParser("Greedy planner (so-planner)")
    p.add_argument("--plan", default="plan of sales.xlsx", help="РџСѓС‚СЊ Рє plan of sales.xlsx")
    p.add_argument("--bom", default="BOM.xlsx", help="РџСѓС‚СЊ Рє BOM.xlsx")
    p.add_argument("--machines", default="machines.xlsx", help="РџСѓС‚СЊ Рє machines.xlsx")
    p.add_argument("--out", default="schedule_out.xlsx", help="РљСѓРґР° СЃРѕС…СЂР°РЅРёС‚СЊ Excel СЃ РїР»Р°РЅРѕРј")
    p.add_argument("--stock", default=None, help="РџСѓС‚СЊ Рє Excel СЃ РѕСЃС‚Р°С‚РєР°РјРё (article/item_id + qty)")
    p.add_argument("--start", default=None, help="РЎС‚Р°СЂС‚РѕРІР°СЏ РґР°С‚Р° (YYYY-MM-DD)")
    p.add_argument("--overload-pct", type=float, default=0.0, help="Р“Р»РѕР±Р°Р»СЊРЅР°СЏ РїРµСЂРµРіСЂСѓР·РєР° (0..1 РёР»Рё 0..100%)")
    p.add_argument("--split-child-orders", action="store_true", help="РљР°Р¶РґС‹Р№ article РІ РѕС‚РґРµР»СЊРЅС‹Р№ order (<base>:<item>)")
    p.add_argument("--stock", default=None, help="РџСѓС‚СЊ Рє Excel СЃ РѕСЃС‚Р°С‚РєР°РјРё (article/item_id + qty)")

    return p.parse_args()

def main():
    args = _parse_args()
    overload = args.overload_pct
    if overload > 1.0:
        overload = overload / 100.0
    out, _ = run_pipeline(
        args.plan, args.bom, args.machines, args.out,
        stock_path=args.stock,
        start_date=args.start,
        overload_pct=overload,
        split_child_orders=bool(args.split_child_orders),
    )
    print(f"Р“РѕС‚РѕРІРѕ: {out}")


if __name__ == "__main__":
    main()

def run_pipeline(
    plan_path: str | None = None,
    bom_path: str | None = None,
    machines_path: str | None = None,
    out_xlsx: str | None = None,
    stock_path: str | None = None,
    start_date: str | None = None,
    overload_pct: float = 0.0,
    split_child_orders: bool = True,
    align_roots_to_due: bool = True,
    reserved_order_ids: Iterable[str] | None = None,
    fixed_order_qty: dict[str, float] | None = None,
    mode: str = "",
    **kwargs,
):
    """
    End-to-end pipeline.
    mode == 'standard_up' -> include parents in schedule.
    otherwise -> standard pipeline (as before).
    """
    mode = (mode or "").lower().strip()

    def _parse_start(date_str: str | None) -> dt.date | None:
        if not date_str:
            return None
        try:
            return pd.to_datetime(date_str).date()
        except Exception:
            return None

    # --- Standard pipeline ---
    if not plan_path:
        raise ValueError("plan_path is required for standard pipeline")
    plan_df = load_plan_of_sales(Path(plan_path))

    demand = build_demand(plan_df, reserved_order_ids=reserved_order_ids)
    bom = load_bom(Path(bom_path))
    machines = load_machines(Path(machines_path))

    stock_map = None
    if stock_path:
        try:
            _stock_df = load_stock_any(Path(stock_path))
            if _stock_df is not None:
                if "workshop" not in _stock_df.columns:
                    _stock_df["workshop"] = ""
                _stock_df = (
                    _stock_df
                    .assign(item_id=lambda x: x["item_id"].astype(str),
                            workshop=lambda x: x["workshop"].astype(str))
                    .groupby(["item_id", "workshop"], as_index=False)["stock_qty"].sum()
                )
                stock_map = {
                    (str(r.item_id), str(r.workshop)): float(r.stock_qty)
                    for r in _stock_df.itertuples(index=False)
                }
        except Exception as e:
            print("[GREEDY WARN] stock load failed:", e)
            stock_map = None

    sched = greedy_schedule(
        demand,
        bom,
        machines,
        start_date=_parse_start(start_date),
        overload_pct=overload_pct,
        split_child_orders=split_child_orders,
        align_roots_to_due=align_roots_to_due,
        stock_map=stock_map,
        reserved_order_ids=reserved_order_ids,
        fixed_order_qty=fixed_order_qty,
        include_parents=(mode == "standard_up"),
    )

    sched_reb = _rebalance_unfixed_by_item_schedule(
        sched,
        bom,
        stock_map,
        fixed_order_qty,
    )
    if sched_reb is None or sched_reb.empty:
        logging.getLogger("so_planner.rebalance").warning(
            "rebalance produced empty schedule; keeping original"
        )
    else:
        sched = sched_reb

    export_with_charts._machines_df = machines
    out_file = export_with_charts(sched, Path(out_xlsx), bom=bom)
    return out_file, sched


def run_greedy(*args, **kwargs):
    """
    Backward-compatible wrapper for run_pipeline with extra 'mode' passthrough.
    """
    # Support old positional signature
    save_to_plan_id = kwargs.pop("save_to_plan_id", None)
    db = kwargs.pop("db", None) or kwargs.pop("session", None)

    def _looks_like_session(x):
        return hasattr(x, "execute") or (hasattr(x, "add") and hasattr(x, "commit"))

    pos = list(args)
    if pos and _looks_like_session(pos[0]):
        db = db or pos.pop(0)

    ordered_names = [
        "plan_path", "bom_path", "machines_path", "out_xlsx", "stock_path",
        "start_date", "overload_pct", "split_child_orders", "align_roots_to_due",
        "guard_limit_days",
    ]
    for i, name in enumerate(ordered_names):
        if i < len(pos) and name not in kwargs:
            kwargs[name] = pos[i]

    # Defaults
    plan_path          = kwargs.get("plan_path")          or "plan of sales.xlsx"
    bom_path           = kwargs.get("bom_path")           or "BOM.xlsx"
    machines_path      = kwargs.get("machines_path")      or "machines.xlsx"
    out_xlsx           = kwargs.get("out_xlsx")           or "schedule_out.xlsx"
    start_date         = kwargs.get("start_date")         or None
    stock_path         = kwargs.get("stock_path")         or "stock_path.xlsx"
    overload_pct       = float(kwargs.get("overload_pct") or 0.0)
    split_child_orders = bool(kwargs.get("split_child_orders") or False)
    align_roots_to_due = bool(kwargs.get("align_roots_to_due") or False)
    mode               = (kwargs.get("mode") or "").lower().strip()
    reserved_order_ids = kwargs.get("reserved_order_ids") or None
    fixed_order_qty    = kwargs.get("fixed_order_qty")    or None

    out_file, sched = run_pipeline(
        plan_path, bom_path, machines_path, out_xlsx,
        stock_path=stock_path,
        start_date=start_date,
        overload_pct=overload_pct,
        split_child_orders=split_child_orders,
        align_roots_to_due=align_roots_to_due,
        reserved_order_ids=reserved_order_ids,
        fixed_order_qty=fixed_order_qty,
        mode=mode,
    )

    # Optional DB save path omitted here to keep merged file simple; can be re-added if needed.
    return out_file, sched

# Final override: ensure modular loaders are used
load_plan_of_sales = _L_load_plan_of_sales
load_bom = _L_load_bom
load_machines = _L_load_machines
load_stock_any = _L_load_stock_any

# Drop dead/legacy loader symbols from namespace to avoid accidental use
try:
    del _dead_load_plan_of_sales
except Exception:
    pass
try:
    del _dead_load_bom
except Exception:
    pass
try:
    del _dead_load_machines
except Exception:
    pass
try:
    del _dead_load_stock_any
except Exception:
    pass
try:
    del _legacy_load_machines
except Exception:
    pass


def build_demand(plan_df: pd.DataFrame, *, reserved_order_ids: Iterable[str] | None = None) -> pd.DataFrame:  # type: ignore[override]
    """Aggregate demand.

    If a 'customer' column exists in plan_df, group by it and keep it in the
    output; order_id will use customer as prefix when present, otherwise item_id.
    """
    group_keys = ["item_id", "due_date"] + (["customer"] if "customer" in plan_df.columns else [])
    g = plan_df.groupby(group_keys, as_index=False).agg(qty=("qty", "sum"))
    sort_keys = ["due_date"] + (["customer"] if "customer" in g.columns else []) + ["item_id"]
    g = g.sort_values(sort_keys, kind="stable").reset_index(drop=True)

    from collections import defaultdict
    seq: dict[tuple[str, object], int] = defaultdict(int)
    reserved = {str(x) for x in (reserved_order_ids or []) if str(x)}
    used: set[str] = set()
    order_ids: list[str] = []
    for _, r in g.iterrows():
        base = str(r.get("customer", "") or r["item_id"])
        key = (base, r["due_date"])  # sequence per (customer|item, date)
        while True:
            seq[key] = seq.get(key, 0) + 1
            oid = f"{base}-{pd.to_datetime(r['due_date']).strftime('%Y%m%d')}-{seq[key]:04d}"
            if oid in reserved or oid in used:
                continue
            break
        order_ids.append(oid)
        used.add(oid)
    g["order_id"] = order_ids
    g["priority"] = pd.to_datetime(g["due_date"])  # default priority
    cols = ["order_id", "item_id", "due_date", "qty", "priority"]
    if "customer" in g.columns:
        cols.append("customer")
    return g[cols]


def expand_demand_with_hierarchy(
    demand: pd.DataFrame,
    bom: pd.DataFrame,
    *,
    split_child_orders: bool = False,
    include_parents: bool = False,
    reserved_order_ids: Iterable[str] | None = None,
    fixed_order_qty: dict[str, float] | None = None,
    stock_map: dict | None = None,
    item_workshop: dict[str, str] | None = None,
) -> pd.DataFrame:
    # Build parent and children maps from BOM
    parents: dict[str, str] = {}
    children_map: dict[str, dict[str, float]] = {}
    for r in bom.itertuples(index=False):
        p = r.root_item_id
        c = r.item_id
        parents[str(c)] = str(p)
        if p and p != c:
            children_map.setdefault(str(p), {})[str(c)] = float(getattr(r, "qty_per_parent", 1.0)) or 1.0

    lag_map_by_edge: dict[tuple[str, str], int] = {}
    if "lag_days" in bom.columns:
        try:
            tmp = bom[["root_item_id", "item_id", "lag_days"]].copy()
            tmp["root_item_id"] = tmp["root_item_id"].astype(str).str.strip()
            tmp["item_id"] = tmp["item_id"].astype(str).str.strip()
            tmp["lag_days"] = pd.to_numeric(tmp["lag_days"], errors="coerce").fillna(0).astype(int)
            tmp = tmp[(tmp["root_item_id"] != "") & (tmp["root_item_id"] != tmp["item_id"])]
            if not tmp.empty:
                tmp = tmp.groupby(["root_item_id", "item_id"], as_index=False)["lag_days"].max()
                lag_map_by_edge = {
                    (str(r.root_item_id), str(r.item_id)): int(r.lag_days)
                    for r in tmp.itertuples(index=False)
                }
        except Exception:
            lag_map_by_edge = {}

    def ancestors(x: str) -> list[str]:
        out: list[str] = []
        seen: set[str] = set()
        cur = x
        for _ in range(100000):
            p = parents.get(cur, "")
            if not p or p in seen or p == cur:
                break
            out.append(p)
            seen.add(p)
            cur = p
        return out

    reserved = {str(x) for x in (reserved_order_ids or []) if str(x)}
    fixed_qty: dict[str, float] = {}
    for k, v in (fixed_order_qty or {}).items():
        try:
            fv = float(v)
        except Exception:
            continue
        if not np.isfinite(fv):
            continue
        fixed_qty[str(k)] = max(0.0, fv)
    used: set[str] = set()
    base_order_map: dict[str, str] = {}
    order_id_map: dict[tuple[str, str, str], str] = {}
    item_ws = item_workshop or {}

    smap = None
    if stock_map:
        smap = {}
        for k, v in stock_map.items():
            if isinstance(k, tuple) and len(k) == 2:
                smap[(str(k[0]), str(k[1]))] = float(v or 0.0)
            else:
                smap[str(k)] = float(v or 0.0)

    def _consume_stock(item: str, qty: float, wk: str) -> float:
        if not smap or qty <= 0:
            return qty
        needed = float(qty)
        tried = []
        for key in ((item, wk), (item, ""), item):
            if key in tried:
                continue
            tried.append(key)
            avail = smap.get(key, 0.0)
            if avail <= 0:
                continue
            take = min(needed, float(avail))
            smap[key] = float(avail) - take
            needed -= take
            if needed <= 0:
                break
        if needed > 0:
            for key, avail in list(smap.items()):
                if isinstance(key, tuple) and len(key) == 2 and str(key[0]) == item and key not in tried:
                    if avail <= 0:
                        continue
                    take = min(needed, float(avail))
                    smap[key] = float(avail) - take
                    needed -= take
                    if needed <= 0:
                        break
        return needed

    def _shift_due(due, item: str, parent_item_id: str | None = None):
        if due is None or pd.isna(due):
            return due
        try:
            base_due = pd.to_datetime(due).date()
        except Exception:
            return due
        parent_item = str(parent_item_id or "")
        lag = 0
        if parent_item:
            lag = int(lag_map_by_edge.get((parent_item, str(item)), 0) or 0)
        if lag <= 0:
            return base_due
        try:
            return base_due - dt.timedelta(days=lag)
        except Exception:
            return base_due

    def _unique_oid(oid: str) -> str:
        if oid not in reserved and oid not in used:
            used.add(oid)
            return oid
        i = 1
        while True:
            cand = f"{oid}~{i}"
            if cand not in reserved and cand not in used:
                used.add(cand)
                return cand
            i += 1

    def _order_id(base_oid: str, item: str, parent_item_id: str | None = None) -> str:
        if split_child_orders:
            key = (base_oid, item, str(parent_item_id or ""))
            if key not in order_id_map:
                order_id_map[key] = _unique_oid(f"{base_oid}:{item}")
            return order_id_map[key]
        if base_oid not in base_order_map:
            base_order_map[base_oid] = _unique_oid(base_oid)
        return base_order_map[base_oid]

    rows: list[dict[str, object]] = []

    def _add_row(
        base_oid_row: str,
        oid: str,
        item: str,
        qty: float,
        due,
        pr,
        role: str,
        cust,
        parent_item_id: str | None,
    ) -> None:
        qty_int = int(round(qty))
        if qty_int <= 0:
            return
        row = {
            "base_order_id": base_oid_row,
            "order_id": oid,
            "item_id": item,
            "due_date": due,
            "qty": qty_int,
            "priority": pr,
            "role": role,
            "customer": (str(cust) if cust is not None else None),
        }
        if parent_item_id:
            row["parent_item_id"] = str(parent_item_id)
        rows.append(row)

    def _plan_item(
        base_oid: str,
        base_oid_row: str,
        item: str,
        required_qty: float,
        due,
        pr,
        role: str,
        cust,
        parent_item_id: str | None,
        path: set[str],
    ) -> float:
        req = float(required_qty) if np.isfinite(required_qty) else 0.0
        req = max(req, 0.0)
        oid_fixed = _order_id(base_oid, item, parent_item_id)
        fixed = float(fixed_qty.get(oid_fixed, 0.0) or 0.0)
        if fixed < 0:
            fixed = 0.0
        target = max(req, fixed)
        extra = max(target - fixed, 0.0)
        wk = item_ws.get(item, "")
        remaining = _consume_stock(item, extra, wk)
        new_qty = remaining
        production = fixed + new_qty

        if fixed > 0:
            _add_row(base_oid_row, oid_fixed, item, fixed, due, pr, role, cust, parent_item_id)
        if new_qty > 0:
            oid_new = _unique_oid(oid_fixed) if fixed > 0 else oid_fixed
            _add_row(base_oid_row, oid_new, item, new_qty, due, pr, role, cust, parent_item_id)

        if production <= 0:
            return 0.0

        for ch, mult in (children_map.get(item, {}) or {}).items():
            ch = str(ch)
            if ch in path:
                continue
            mult_val = mult if np.isfinite(mult) and mult > 0 else 1.0
            child_req = production * float(mult_val)
            child_due = _shift_due(due, ch, parent_item_id=item)
            _plan_item(
                base_oid,
                base_oid_row,
                ch,
                child_req,
                child_due,
                pr,
                "CHILD",
                cust,
                item,
                path | {ch},
            )
        return production

    for r in demand.itertuples(index=False):
        base_oid = str(r.order_id)
        it = str(r.item_id)
        due = r.due_date
        qty = float(r.qty) if not pd.isna(r.qty) else 0.0
        pr = r.priority
        cust = getattr(r, "customer", None)
        base_oid_row = base_oid if split_child_orders else _order_id(base_oid, it)
        prod_qty = _plan_item(
            base_oid,
            base_oid_row,
            it,
            qty,
            due,
            pr,
            "FG",
            cust,
            None,
            {it},
        )
        if include_parents and prod_qty > 0:
            for a in ancestors(it):
                oid_parent = _unique_oid(_order_id(base_oid, a))
                _add_row(base_oid_row, oid_parent, a, prod_qty, due, pr, "PARENT", cust, None)

    exp = pd.DataFrame(rows)
    if exp.empty:
        raise ValueError("Greedy: expanded_demand is empty (check BOM/qty_per_parent)")
    links = build_bom_hierarchy(bom)
    depth_map = {r.item_id: int(r.level) for r in links.itertuples(index=False)} if not links.empty else {}
    exp["depth"] = exp["item_id"].map(depth_map).fillna(0).astype(int)
    exp = exp.sort_values(["priority", "depth", "item_id"], kind="stable").reset_index(drop=True)
    return exp



